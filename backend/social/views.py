from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings
from django.http import HttpResponse
from io import BytesIO
import os
import subprocess
import sys
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import datetime
from .models import ApiLog, Channel, Post, DailySnapshot, FollowerSnapshot
from authentication.permissions import IsAuthenticated, IsManagerOrAdmin
from .sync import SyncEngine

def get_today_date():
    return timezone.now().strftime('%Y-%m-%d')

def get_recent_start_date(days):
    return (timezone.now() - datetime.timedelta(days=days)).strftime('%Y-%m-%d')

def resolve_reporting_period(query):
    start_date = query.get('startDate')
    end_date = query.get('endDate')
    
    if start_date and end_date:
        return start_date, end_date
    return get_recent_start_date(6), get_today_date()

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def channels_list(request):
    if request.method == 'GET':
        channels = Channel.objects.exclude(external_id='current-facebook-token').order_by('name')
        
        # Calculate post count for each channel
        post_counts = {}
        for post in Post.objects.all():
            post_counts[post.channel_id] = post_counts.get(post.channel_id, 0) + 1
            
        result = []
        for c in channels:
            result.append({
                "id": c.id,
                "platform": c.platform,
                "name": c.name,
                "externalId": c.external_id,
                "status": c.status,
                "timezone": c.timezone,
                "createdAt": c.created_at.isoformat(),
                "updatedAt": c.updated_at.isoformat(),
                "lastSyncAt": c.last_sync_at.isoformat() if c.last_sync_at else None,
                "lastSyncStatus": c.last_sync_status,
                "followersCount": c.followers_count,
                "totalPosts": post_counts.get(c.id, 0)
            })
        return Response(result)
        
    elif request.method == 'POST':
        # Add new channel
        if getattr(request, 'user_role', 'EMPLOYEE') not in ['ADMIN', 'MANAGER']:
            return Response({"error": "Quyền truy cập bị từ chối."}, status=status.HTTP_403_FORBIDDEN)
            
        data = request.data or {}
        channel_id = data.get('id')
        platform = data.get('platform')
        name = data.get('name')
        external_id = data.get('externalId') or data.get('external_id')
        
        if not channel_id or not platform or not name or not external_id:
            return Response({"error": "Thiếu thông tin kênh."}, status=status.HTTP_400_BAD_REQUEST)
            
        channel, created = Channel.objects.update_or_create(
            id=channel_id,
            defaults={
                'platform': platform,
                'name': name,
                'external_id': external_id,
                'status': data.get('status', 'active'),
                'timezone': data.get('timezone', 'Asia/Bangkok'),
                'created_at': timezone.now(),
                'updated_at': timezone.now()
            }
        )
        return Response({
            "id": channel.id,
            "name": channel.name,
            "platform": channel.platform,
            "externalId": channel.external_id
        })

@api_view(['PUT', 'DELETE'])
@permission_classes([IsManagerOrAdmin])
def channel_detail(request, channel_id):
    try:
        channel = Channel.objects.get(id=channel_id)
    except Channel.DoesNotExist:
        return Response({"error": "Không tìm thấy kênh."}, status=status.HTTP_404_NOT_FOUND)
        
    if request.method == 'PUT':
        data = request.data or {}
        if 'name' in data:
            channel.name = data['name']
        if 'externalId' in data or 'external_id' in data:
            channel.external_id = data.get('externalId') or data.get('external_id')
        if 'status' in data:
            channel.status = data['status']
        if 'timezone' in data:
            channel.timezone = data['timezone']
            
        channel.updated_at = timezone.now()
        channel.save()
        return Response({
            "id": channel.id,
            "name": channel.name,
            "platform": channel.platform,
            "externalId": channel.external_id
        })
        
    elif request.method == 'DELETE':
        # Delete channel and associated posts and snapshots
        channel.delete()
        Post.objects.filter(channel_id=channel_id).delete()
        DailySnapshot.objects.filter(channel_id=channel_id).delete()
        FollowerSnapshot.objects.filter(channel_id=channel_id).delete()
        return Response({"success": True, "id": channel_id})

@api_view(['POST'])
@permission_classes([IsManagerOrAdmin])
def channel_test_connection(request, channel_id):
    try:
        channel = Channel.objects.get(id=channel_id)
        provider = SyncEngine.get_provider(channel.platform)
        valid = provider.validate_credentials(channel.id, channel.external_id)
        return Response({"success": valid})
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsManagerOrAdmin])
def channel_sync(request, channel_id):
    since = request.data.get('since')
    until = request.data.get('until')
    success, message = SyncEngine.sync_channel(channel_id, since=since, until=until)
    if success:
        return Response({"success": True, "message": message})
    return Response({"error": message}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def media_summary_trend(request):
    group_by = request.query_params.get('groupBy', 'month')
    platform_filter = request.query_params.get('platform')
    channel_id_filter = request.query_params.get('channelId')
    
    if platform_filter == 'all':
        platform_filter = None
    if channel_id_filter == 'all':
        channel_id_filter = None
        
    # Build list of active channels
    channels = Channel.objects.filter(status='active')
    if platform_filter:
        channels = channels.filter(platform=platform_filter)
    if channel_id_filter:
        channels = channels.filter(id=channel_id_filter)
        
    channel_ids = [c.id for c in channels]
    
    # Calculate buckets
    now = timezone.now()
    buckets = []
    
    # We want 13 months, 8 quarters, or 5 years
    if group_by == 'month':
        for i in range(12, -1, -1):
            # Calculate offset month
            year = now.year
            month = now.month - i
            while month <= 0:
                month += 12
                year -= 1
            # start/end of that month
            start_date = f"{year}-{month:02d}-01"
            # end date: get last day
            if month == 12:
                end_date = f"{year}-12-31"
            else:
                next_month = datetime.date(year, month + 1, 1)
                last_day = next_month - datetime.timedelta(days=1)
                end_date = last_day.strftime('%Y-%m-%d')
                
            buckets.append({
                'key': f"{year}-{month:02d}",
                'label': f"T{month}/{year}",
                'start': start_date,
                'end': end_date
            })
    elif group_by == 'quarter':
        for i in range(7, -1, -1):
            curr_q = (now.month - 1) // 3
            # offset quarter
            q_offset = curr_q - i
            year = now.year
            while q_offset < 0:
                q_offset += 4
                year -= 1
            q_num = q_offset + 1
            start_month = q_offset * 3 + 1
            start_date = f"{year}-{start_month:02d}-01"
            end_month = start_month + 2
            if end_month == 12:
                end_date = f"{year}-12-31"
            else:
                next_month = datetime.date(year, end_month + 1, 1)
                last_day = next_month - datetime.timedelta(days=1)
                end_date = last_day.strftime('%Y-%m-%d')
                
            buckets.append({
                'key': f"{year}-Q{q_num}",
                'label': f"Q{q_num}/{year}",
                'start': start_date,
                'end': end_date
            })
    else: # year
        for i in range(4, -1, -1):
            y = now.year - i
            buckets.append({
                'key': str(y),
                'label': str(y),
                'start': f"{y}-01-01",
                'end': f"{y}-12-31"
            })
            
    period_start = buckets[0]['start']
    period_end = buckets[-1]['end']
    
    # Query posts
    posts = Post.objects.filter(
        channel_id__in=channel_ids,
        published_at__gte=period_start,
        published_at__lte=f"{period_end}T23:59:59.999Z"
    )
    
    # Query snapshots
    snapshots = DailySnapshot.objects.filter(
        channel_id__in=channel_ids,
        snapshot_date__gte=period_start,
        snapshot_date__lte=period_end
    )
    
    # Get latest snapshot for each postKey
    latest_snaps = {}
    for snap in snapshots:
        existing = latest_snaps.get(snap.post_key)
        if not existing or snap.snapshot_date > existing.snapshot_date:
            latest_snaps[snap.post_key] = snap
            
    # Query follower snapshots
    follower_snaps = FollowerSnapshot.objects.filter(
        channel_id__in=channel_ids,
        snapshot_date__gte=period_start,
        snapshot_date__lte=period_end
    ).order_by('snapshot_date')
    
    # Process follower snapshot trend by bucket
    follower_by_bucket = {}
    latest_followers = {}
    f_index = 0
    follower_list = list(follower_snaps)
    
    for bucket in buckets:
        while f_index < len(follower_list) and follower_list[f_index].snapshot_date <= bucket['end']:
            snap = follower_list[f_index]
            latest_followers[snap.channel_id] = snap.followers_count
            f_index += 1
        follower_by_bucket[bucket['key']] = dict(latest_followers)
        
    trend = []
    for bucket in buckets:
        b_start = bucket['start']
        b_end = bucket['end']
        
        # Filter posts in this bucket range
        b_posts = [p for p in posts if str(p.published_at)[:10] >= b_start and str(p.published_at)[:10] <= b_end]
        
        views_sum = 0
        engagement_sum = 0
        
        for p in b_posts:
            snap = latest_snaps.get(p.post_key)
            if snap:
                views_sum += getattr(snap, 'views', 0) or getattr(snap, 'impressions', 0) or getattr(snap, 'reach', 0) or 0
                engagement_sum += getattr(snap, 'total_engagement', 0) or 0
                
        # Followers count
        followers_map = follower_by_bucket.get(bucket['key'], {})
        followers_sum = sum(followers_map.values())
        
        trend.append({
            'period': bucket['key'],
            'label': bucket['label'],
            'views': views_sum,
            'engagement': engagement_sum,
            'postsCount': len(b_posts),
            'followers': followers_sum
        })
        
    return Response({'groupBy': group_by, 'trend': trend})

def _build_media_summary(query):
    period_start, period_end = resolve_reporting_period(query)
    platform_filter = query.get('platform')
    channel_id_filter = query.get('channelId')

    if platform_filter == 'all':
        platform_filter = None
    if channel_id_filter == 'all':
        channel_id_filter = None

    channels = Channel.objects.filter(status='active')
    if platform_filter:
        channels = channels.filter(platform=platform_filter)
    if channel_id_filter:
        channels = channels.filter(id=channel_id_filter)

    channel_ids = [channel.id for channel in channels]
    period_start_dt = SyncEngine._parse_boundary(period_start)
    period_end_dt = SyncEngine._parse_boundary(period_end, end_of_day=True)
    posts = Post.objects.filter(
        channel_id__in=channel_ids,
        published_at__gte=period_start_dt,
        published_at__lte=period_end_dt,
    )
    snapshots = DailySnapshot.objects.filter(
        channel_id__in=channel_ids,
        snapshot_date__gte=period_start,
        snapshot_date__lte=period_end,
    )

    latest_snapshots = {}
    for snapshot in snapshots:
        existing = latest_snapshots.get(snapshot.post_key)
        if not existing or snapshot.snapshot_date > existing.snapshot_date:
            latest_snapshots[snapshot.post_key] = snapshot

    result = []
    for channel in channels:
        channel_posts = [post for post in posts if post.channel_id == channel.id]
        reactions = comments = shares = clicks = views = reach = impressions = 0

        for post in channel_posts:
            snapshot = latest_snapshots.get(post.post_key)
            if not snapshot:
                continue
            reactions += snapshot.reactions
            comments += snapshot.comments
            shares += snapshot.shares
            clicks += snapshot.clicks
            views += snapshot.views or snapshot.impressions or snapshot.reach or 0
            reach += snapshot.reach
            impressions += snapshot.impressions

        total_engagement = reactions + comments + shares + clicks
        engagement_rate = None
        if reach > 0:
            engagement_rate = round((total_engagement / reach) * 100, 2)
        elif impressions > 0:
            engagement_rate = round((total_engagement / impressions) * 100, 2)

        result.append({
            'id': channel.id,
            'channelId': channel.id,
            'name': channel.name,
            'channelName': channel.name,
            'externalId': channel.external_id,
            'platform': channel.platform,
            'lastSyncAt': channel.last_sync_at.isoformat() if channel.last_sync_at else None,
            'lastSyncStatus': channel.last_sync_status,
            'postsCount': len(channel_posts),
            'followersCount': channel.followers_count,
            'views': views,
            'reactions': reactions,
            'comments': comments,
            'shares': shares,
            'clicks': clicks,
            'totalEngagement': total_engagement,
            'engagementRate': engagement_rate,
        })

    return result


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def media_summary(request):
    return Response(_build_media_summary(request.query_params))


def _safe_excel_value(value):
    if isinstance(value, str) and value.startswith(('=', '+', '-', '@')):
        return "'" + value
    return value


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def media_summary_xlsx(request):
    rows = _build_media_summary(request.query_params)
    period_start, period_end = resolve_reporting_period(request.query_params)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Bao cao tong hop'
    sheet.append(['BÁO CÁO TỔNG HỢP TRUYỀN THÔNG'])
    sheet.append(['Thời gian', period_start + ' - ' + period_end])
    sheet.append([])

    headers = [
        'STT',
        'Nền tảng',
        'Tên trang',
        'ID trang',
        'Đồng bộ lần cuối',
        'Người theo dõi',
        'Số bài đăng',
        'Lượt xem',
        'Cảm xúc',
        'Bình luận',
        'Chia sẻ',
        'Lượt nhấp',
        'Tổng tương tác',
        'Tỷ lệ tương tác (%)',
    ]
    sheet.append(headers)

    for index, row in enumerate(rows, start=1):
        platform_name = 'Facebook' if row['platform'] == 'facebook' else 'Zalo OA'
        sheet.append([
            index,
            platform_name,
            _safe_excel_value(row['name']),
            _safe_excel_value(row['externalId']),
            row['lastSyncAt'] or '',
            row['followersCount'],
            row['postsCount'],
            row['views'],
            row['reactions'],
            row['comments'],
            row['shares'],
            row['clicks'],
            row['totalEngagement'],
            row['engagementRate'] if row['engagementRate'] is not None else '',
        ])

    title_fill = PatternFill('solid', fgColor='17365D')
    header_fill = PatternFill('solid', fgColor='2563EB')
    sheet['A1'].fill = title_fill
    sheet['A1'].font = Font(color='FFFFFF', bold=True, size=14)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for cell in sheet[4]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.freeze_panes = 'A5'
    sheet.auto_filter.ref = f"A4:N{max(4, sheet.max_row)}"
    widths = [7, 14, 34, 24, 23, 18, 16, 16, 14, 14, 14, 14, 20, 24]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + column)].width = width

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="bao_cao_tong_hop_{get_today_date()}.xlsx"'
    )
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def followers_trend(request):
    try:
        if request.query_params.get('startDate') and request.query_params.get('endDate'):
            period_start, period_end = resolve_reporting_period(request.query_params)
        else:
            raw_days = int(request.query_params.get('days', 7))
            if not (1 <= raw_days <= 365):
                raise ValueError
            period_start = get_recent_start_date(raw_days - 1)
            period_end = get_today_date()
        start_day = datetime.date.fromisoformat(period_start)
        end_day = datetime.date.fromisoformat(period_end)
        if start_day > end_day or (end_day - start_day).days > 365:
            raise ValueError
    except (TypeError, ValueError):
        return Response(
            {"error": "Kho\u1ea3ng th\u1eddi gian ph\u1ea3i h\u1ee3p l\u1ec7 v\u00e0 kh\u00f4ng v\u01b0\u1ee3t qu\u00e1 1 n\u0103m."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    channel_id_filter = request.query_params.get('channelId')
    platform_filter = request.query_params.get('platform')
    if channel_id_filter == 'all':
        channel_id_filter = None
    if platform_filter == 'all':
        platform_filter = None

    channels = Channel.objects.filter(status='active').exclude(external_id='current-facebook-token')
    if platform_filter:
        channels = channels.filter(platform=platform_filter)
    if channel_id_filter:
        channels = channels.filter(id=channel_id_filter)
    channel_ids = list(channels.values_list('id', flat=True))

    histories = {channel_id: [] for channel_id in channel_ids}
    snapshots_by_day = {channel_id: {} for channel_id in channel_ids}
    snapshots = FollowerSnapshot.objects.filter(
        channel_id__in=channel_ids,
        snapshot_date__lte=period_end,
    ).order_by('snapshot_date')
    for snapshot in snapshots:
        histories.setdefault(snapshot.channel_id, []).append(snapshot)
        snapshots_by_day.setdefault(snapshot.channel_id, {})[snapshot.snapshot_date] = snapshot

    cursors = {channel_id: 0 for channel_id in channel_ids}
    latest_counts = {}
    trend = []
    current_day = start_day
    while current_day <= end_day:
        date_string = current_day.isoformat()
        daily_follows = []
        daily_unfollows = []
        for channel_id in channel_ids:
            history = histories.get(channel_id, [])
            cursor = cursors[channel_id]
            while cursor < len(history) and history[cursor].snapshot_date <= date_string:
                latest_counts[channel_id] = history[cursor].followers_count
                cursor += 1
            cursors[channel_id] = cursor

            daily_snapshot = snapshots_by_day.get(channel_id, {}).get(date_string)
            if daily_snapshot and daily_snapshot.daily_follows_unique is not None:
                daily_follows.append(daily_snapshot.daily_follows_unique)
            if daily_snapshot and daily_snapshot.daily_unfollows_unique is not None:
                daily_unfollows.append(daily_snapshot.daily_unfollows_unique)

        trend.append({
            'date': date_string,
            'followersCount': sum(latest_counts.get(channel_id, 0) for channel_id in channel_ids),
            'dailyFollowsUnique': sum(daily_follows) if daily_follows else None,
            'dailyUnfollowsUnique': sum(daily_unfollows) if daily_unfollows else None,
        })
        current_day += datetime.timedelta(days=1)

    return Response(trend)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_view(request):
    platform_filter = request.query_params.get('platform')
    channel_id_filter = request.query_params.get('channelId')
    post_type_filter = request.query_params.get('postType')
    start_date = request.query_params.get('startDate')
    end_date = request.query_params.get('endDate')
    
    if platform_filter == 'all':
        platform_filter = None
    if channel_id_filter == 'all':
        channel_id_filter = None
    if post_type_filter == 'all':
        post_type_filter = None
        
    period_start, period_end = resolve_reporting_period({'startDate': start_date, 'endDate': end_date})
    
    # 1. Fetch active channels
    channels = Channel.objects.filter(status='active').exclude(external_id='current-facebook-token')
    if platform_filter:
        channels = channels.filter(platform=platform_filter)
    if channel_id_filter:
        channels = channels.filter(id=channel_id_filter)
        
    channel_ids = [c.id for c in channels]
    
    # 2. Fetch posts
    posts = Post.objects.filter(channel_id__in=channel_ids)
    if platform_filter:
        posts = posts.filter(platform=platform_filter)
    if post_type_filter:
        if post_type_filter == 'other':
            posts = posts.exclude(post_type__in=['photo', 'video', 'link', 'status'])
        else:
            posts = posts.filter(post_type=post_type_filter)
        
    # Filter by date window
    posts = [p for p in posts if period_start <= str(p.published_at)[:10] <= period_end]
    post_keys = [p.post_key for p in posts]
    
    # 3. Fetch snapshots
    snapshots = DailySnapshot.objects.filter(
        channel_id__in=channel_ids,
        snapshot_date__gte=period_start,
        snapshot_date__lte=period_end
    )
    if platform_filter:
        snapshots = snapshots.filter(platform=platform_filter)
        
    # Get latest snapshots
    latest_snaps = {}
    for snap in snapshots:
        if snap.post_key not in post_keys:
            continue
        existing = latest_snaps.get(snap.post_key)
        if not existing or snap.snapshot_date > existing.snapshot_date:
            latest_snaps[snap.post_key] = snap
            
    # Follower is a stock metric: use the last saved value on or before the selected end date.
    follower_snapshots = FollowerSnapshot.objects.filter(
        channel_id__in=channel_ids,
        snapshot_date__lte=period_end,
    ).order_by('snapshot_date')
    latest_follower_snapshots = {}
    for snapshot in follower_snapshots:
        latest_follower_snapshots[snapshot.channel_id] = snapshot
    followers_count = sum(snapshot.followers_count for snapshot in latest_follower_snapshots.values())
    followers_available = bool(latest_follower_snapshots)

    # Calculate KPIs
    posts_count = len(posts)
    reactions = 0
    comments = 0
    shares = 0
    views = 0
    reach = 0
    impressions = 0
    clicks = 0
    
    for snap in latest_snaps.values():
        reactions += snap.reactions
        comments += snap.comments
        shares += snap.shares
        views += snap.views or snap.impressions or snap.reach or 0
        reach += snap.reach
        impressions += snap.impressions
        clicks += snap.clicks
        
    total_engagement = reactions + comments + shares + clicks
    engagement_rate = None
    if reach > 0:
        engagement_rate = round((total_engagement / reach) * 100, 2)
    elif impressions > 0:
        engagement_rate = round((total_engagement / impressions) * 100, 2)
        
    # 4. Generate Trends
    channel_map = {c.id: c.name for c in channels}
    trend_map = {}
    
    for p in posts:
        date_str = str(p.published_at)[:10]
        snap = latest_snaps.get(p.post_key)
        chan_name = channel_map.get(p.channel_id, 'Kênh ẩn')
        
        curr = trend_map.get(date_str) or {
            'date': date_str,
            'engagement': 0,
            'postsCount': 0,
            'likes': 0,
            'comments': 0,
            'shares': 0,
            'views': 0,
            'reach': 0
        }
        
        v_count = (snap.views or snap.impressions or snap.reach or 0) if snap else 0
        eng_count = snap.total_engagement if snap else 0
        likes_count = snap.reactions if snap else 0
        comments_count = snap.comments if snap else 0
        shares_count = snap.shares if snap else 0
        reach_count = snap.reach if snap else 0
        
        curr['engagement'] += eng_count
        curr['postsCount'] += 1
        curr['likes'] += likes_count
        curr['comments'] += comments_count
        curr['shares'] += shares_count
        curr['views'] += v_count
        curr['reach'] += reach_count
        
        # Add channel metrics
        curr[f"{chan_name}_engagement"] = curr.get(f"{chan_name}_engagement", 0) + eng_count
        curr[f"{chan_name}_postsCount"] = curr.get(f"{chan_name}_postsCount", 0) + 1
        curr[f"{chan_name}_likes"] = curr.get(f"{chan_name}_likes", 0) + likes_count
        curr[f"{chan_name}_comments"] = curr.get(f"{chan_name}_comments", 0) + comments_count
        curr[f"{chan_name}_shares"] = curr.get(f"{chan_name}_shares", 0) + shares_count
        curr[f"{chan_name}_views"] = curr.get(f"{chan_name}_views", 0) + v_count
        curr[f"{chan_name}_reach"] = curr.get(f"{chan_name}_reach", 0) + reach_count
        
        trend_map[date_str] = curr
        
    trends = []
    for point in trend_map.values():
        for c in channels:
            metrics = ['engagement', 'postsCount', 'likes', 'comments', 'shares', 'views', 'reach']
            for m in metrics:
                key = f"{c.name}_{m}"
                if key not in point:
                    point[key] = 0
                    
        point['engagementRate'] = round((point['engagement'] / point['reach']) * 100, 2) if point['reach'] > 0 else 0
        for c in channels:
            c_eng = point.get(f"{c.name}_engagement", 0)
            c_reach = point.get(f"{c.name}_reach", 0)
            point[f"{c.name}_engagementRate"] = round((c_eng / c_reach) * 100, 2) if c_reach > 0 else 0
            
        trends.append(point)
    trends.sort(key=lambda x: x['date'])
    
    # 5. Channel Stats
    channel_stats = []
    for c in channels:
        c_posts = [p for p in posts if p.channel_id == c.id]
        c_eng = sum(latest_snaps.get(p.post_key).total_engagement for p in c_posts if latest_snaps.get(p.post_key))
        channel_stats.append({
            'channelName': c.name,
            'platform': c.platform,
            'postsCount': len(c_posts),
            'engagement': c_eng
        })
        
    # 6. Top Posts
    top_posts = []
    for p in posts:
        snap = latest_snaps.get(p.post_key)
        top_posts.append({
            'postKey': p.post_key,
            'platform': p.platform,
            'channelId': p.channel_id,
            'externalPostId': p.external_post_id,
            'postUrl': p.post_url,
            'imageUrl': p.image_url,
            'postType': p.post_type,
            'message': p.message,
            'publishedAt': p.published_at,
            'importedAt': p.imported_at,
            'updatedAt': p.updated_at,
            'isDeleted': p.is_deleted,
            'engagement': snap.total_engagement if snap else 0,
            'likes': snap.reactions if snap else 0,
            'comments': snap.comments if snap else 0,
            'shares': snap.shares if snap else 0,
            'views': (snap.views or snap.impressions or snap.reach or 0) if snap else 0
        })
    top_posts.sort(key=lambda x: x['publishedAt'], reverse=True)
    top_posts = top_posts[:10]
    
    # 7. Top Viewed Posts (past 12 months)
    top_viewed_end_day = timezone.localdate()
    top_viewed_start_day = top_viewed_end_day - datetime.timedelta(days=365)
    top_viewed_start = top_viewed_start_day.isoformat()
    top_viewed_end = top_viewed_end_day.isoformat()
    top_viewed_start_at = timezone.make_aware(datetime.datetime.combine(top_viewed_start_day, datetime.time.min))
    top_viewed_end_at = timezone.make_aware(datetime.datetime.combine(top_viewed_end_day, datetime.time.max))

    tv_posts = Post.objects.filter(
        channel_id__in=channel_ids,
        published_at__gte=top_viewed_start_at,
        published_at__lte=top_viewed_end_at,
    )
    if post_type_filter:
        if post_type_filter == 'other':
            tv_posts = tv_posts.exclude(post_type__in=['photo', 'video', 'link', 'status'])
        else:
            tv_posts = tv_posts.filter(post_type=post_type_filter)
        
    tv_post_keys = [p.post_key for p in tv_posts]
    
    tv_snapshots = DailySnapshot.objects.filter(
        channel_id__in=channel_ids,
        snapshot_date__gte=top_viewed_start,
        snapshot_date__lte=top_viewed_end
    )
    
    tv_latest_snaps = {}
    for snap in tv_snapshots:
        if snap.post_key not in tv_post_keys:
            continue
        existing = tv_latest_snaps.get(snap.post_key)
        if not existing or snap.snapshot_date > existing.snapshot_date:
            tv_latest_snaps[snap.post_key] = snap
            
    top_viewed_candidates = []
    for p in tv_posts:
        snap = tv_latest_snaps.get(p.post_key)
        v_count = (snap.views or snap.impressions or snap.reach or 0) if snap else 0
        top_viewed_candidates.append({
            'postKey': p.post_key,
            'platform': p.platform,
            'channelId': p.channel_id,
            'externalPostId': p.external_post_id,
            'postUrl': p.post_url,
            'imageUrl': p.image_url,
            'postType': p.post_type,
            'message': p.message,
            'publishedAt': p.published_at,
            'importedAt': p.imported_at,
            'updatedAt': p.updated_at,
            'isDeleted': p.is_deleted,
            'engagement': snap.total_engagement if snap else 0,
            'likes': snap.reactions if snap else 0,
            'comments': snap.comments if snap else 0,
            'shares': snap.shares if snap else 0,
            'views': v_count
        })
    top_viewed_candidates.sort(key=lambda x: x['views'], reverse=True)
    top_viewed_posts = top_viewed_candidates[:5]
    
    # 8. Last Sync At
    last_sync_channel = channels.filter(last_sync_at__isnull=False).order_by('-last_sync_at').first()
    
    errors = []
    for c in channels:
        if c.last_sync_status == 'failed':
            errors.append(f"Kênh \"{c.name}\" ({c.platform.upper()}) bị lỗi token hoặc kết nối.")
            
    # 9. Type Stats
    type_stats_map = {}
    for p in posts:
        raw_type = p.post_type or 'Khác'
        mapped_type = ('Ảnh / Album' if raw_type.lower() in ('photo', 'album')
                       else 'Video / Reel' if raw_type.lower() in ('video', 'reel')
                       else 'Liên kết' if raw_type.lower() == 'link'
                       else 'Văn bản' if raw_type.lower() == 'status'
                       else 'Khác')
        snap = latest_snaps.get(p.post_key)
        v_count = (snap.views or snap.impressions or snap.reach or 0) if snap else 0
        eng = snap.total_engagement if snap else 0
        
        curr = type_stats_map.get(mapped_type) or {
            'type': mapped_type,
            'count': 0,
            'views': 0,
            'engagement': 0,
            'engagementRate': None
        }
        curr['count'] += 1
        curr['views'] += v_count
        curr['engagement'] += eng
        type_stats_map[mapped_type] = curr
        
    type_stats = []
    for stat in type_stats_map.values():
        stat['engagementRate'] = round((stat['engagement'] / stat['views']) * 100, 2) if stat['views'] > 0 else None
        type_stats.append(stat)
        
    # 10. Platform Stats
    platform_stats_map = {}
    for p in posts:
        raw_plat = p.platform or 'facebook'
        mapped_plat = 'Facebook' if raw_plat.lower() == 'facebook' else 'Zalo OA'
        snap = latest_snaps.get(p.post_key)
        eng = snap.total_engagement if snap else 0
        
        curr = platform_stats_map.get(mapped_plat) or {
            'platform': mapped_plat,
            'count': 0,
            'engagement': 0
        }
        curr['count'] += 1
        curr['engagement'] += eng
        platform_stats_map[mapped_plat] = curr
        
    platform_stats = list(platform_stats_map.values())
    
    return Response({
        'kpis': {
            'postsCount': posts_count,
            'reactions': reactions,
            'comments': comments,
            'shares': shares,
            'views': views,
            'reach': reach,
            'totalEngagement': total_engagement,
            'engagementRate': engagement_rate,
            'followers': followers_count,
            'followersAvailable': followers_available
        },
        'trends': trends,
        'channelStats': channel_stats,
        'topPosts': top_posts,
        'topViewedPosts': top_viewed_posts,
        'typeStats': type_stats,
        'platformStats': platform_stats,
        'lastSync': last_sync_channel.last_sync_at.isoformat() if last_sync_channel else None,
        'errors': errors
    })

def _start_background_sync(days=365):
    days = max(1, min(int(days), 3650))
    manage_py = settings.BASE_DIR / "manage.py"
    process_args = [
        sys.executable,
        str(manage_py),
        "sync_social_daily",
        "--backfill-days",
        str(days),
        "--recent-days",
        str(days),
    ]
    popen_options = {
        "cwd": str(settings.BASE_DIR),
        "stdout": subprocess.DEVNULL,
        "stderr": subprocess.DEVNULL,
        "close_fds": True,
    }
    if os.name == "nt":
        popen_options["creationflags"] = (
            subprocess.CREATE_NEW_PROCESS_GROUP | subprocess.DETACHED_PROCESS
        )
    else:
        popen_options["start_new_session"] = True
    subprocess.Popen(process_args, **popen_options)


@api_view(['POST'])
@permission_classes([IsManagerOrAdmin])
def sync_all(request):
    """POST /api/sync/all - Sync all active channels."""
    try:
        if request.data.get("background"):
            running_since = timezone.now() - datetime.timedelta(hours=6)
            if ApiLog.objects.filter(
                status__in=["queued", "running"],
                started_at__gte=running_since,
            ).exists():
                return Response({
                    "success": True,
                    "queued": True,
                    "alreadyRunning": True,
                    "message": "Đang có một lần đồng bộ chạy ngầm. Dữ liệu sẽ tự cập nhật khi hoàn tất.",
                }, status=status.HTTP_202_ACCEPTED)

            days = request.data.get("days", 365)
            _start_background_sync(days)
            return Response({
                "success": True,
                "queued": True,
                "message": "Đã bắt đầu đồng bộ ngầm dữ liệu 1 năm. Bạn có thể tiếp tục sử dụng trang.",
            }, status=status.HTTP_202_ACCEPTED)

        google_token = getattr(request, 'google_access_token', None)
        since = request.data.get('since')
        until = request.data.get('until')
        results = SyncEngine.sync_all_channels(
            google_token,
            since=since,
            until=until,
        )
        return Response(results)
    except (TypeError, ValueError):
        return Response(
            {"error": "Khoảng thời gian đồng bộ không hợp lệ."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    except Exception as exc:
        return Response(
            {"error": str(exc)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def sync_history(request):
    """GET /api/sync/history - Lịch sử đồng bộ"""
    from .models import ApiLog
    logs = ApiLog.objects.all().order_by('-started_at')[:100]
    result = []
    for log in logs:
        result.append({
            'id': log.log_id,
            'logId': log.log_id,
            'startedAt': log.started_at.isoformat(),
            'endedAt': log.ended_at.isoformat() if log.ended_at else None,
            'platform': log.platform,
            'action': log.action,
            'channelId': log.channel_id,
            'status': log.status,
            'recordsReceived': log.records_received,
            'recordsInserted': log.records_inserted,
            'recordsUpdated': log.records_updated,
            'errorCode': log.error_code,
            'errorMessage': log.error_message,
            'requestId': log.request_id
        })
    return Response(result)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def posts_list(request):
    """GET /api/posts - Danh sách bài viết có phân trang"""
    try:
        page = int(request.query_params.get('page', 1))
        limit = int(request.query_params.get('limit', 20))
        start_date = request.query_params.get('startDate', get_recent_start_date(30))
        end_date = request.query_params.get('endDate', get_today_date())
        platform_filter = request.query_params.get('platform')
        channel_id_filter = request.query_params.get('channelId')
        post_type_filter = request.query_params.get('postType')
        sort_by = request.query_params.get('sortBy', 'publishedAt')
        sort_dir = request.query_params.get('sortDir', 'desc')
        
        queryset = Post.objects.filter(is_deleted=False)
        if start_date:
            queryset = queryset.filter(published_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(published_at__date__lte=end_date)
        if platform_filter:
            queryset = queryset.filter(platform=platform_filter)
        if channel_id_filter:
            queryset = queryset.filter(channel_id=channel_id_filter)
        if post_type_filter and post_type_filter != 'all':
            if post_type_filter == 'other':
                queryset = queryset.exclude(post_type__in=['photo', 'video', 'link', 'status'])
            else:
                queryset = queryset.filter(post_type=post_type_filter)
            
        # Active channels only
        active_channel_ids = list(Channel.objects.filter(status='active').values_list('id', flat=True))
        queryset = queryset.filter(channel_id__in=active_channel_ids)
        
        # Sort
        order = '-published_at' if sort_dir == 'desc' else 'published_at'
        if sort_by == 'engagement':
            order = '-published_at'  # Will sort by engagement after snapshot join
        queryset = queryset.order_by(order)
        
        total = queryset.count()
        offset = (page - 1) * limit
        posts = list(queryset[offset:offset + limit])
        
        # Get snapshots for these posts
        post_keys = [p.post_key for p in posts]
        snapshots = DailySnapshot.objects.filter(post_key__in=post_keys)
        latest_snaps = {}
        for s in snapshots:
            existing = latest_snaps.get(s.post_key)
            if not existing or s.snapshot_date > existing.snapshot_date:
                latest_snaps[s.post_key] = s
                
        # Channel name map
        channels = {c.id: c.name for c in Channel.objects.all()}
        
        result = []
        for p in posts:
            snap = latest_snaps.get(p.post_key)
            result.append({
                'postKey': p.post_key,
                'platform': p.platform,
                'channelId': p.channel_id,
                'channelName': channels.get(p.channel_id, ''),
                'externalPostId': p.external_post_id,
                'postUrl': p.post_url,
                'imageUrl': p.image_url,
                'postType': p.post_type,
                'message': p.message or '',
                'publishedAt': p.published_at.isoformat(),
                'reactions': snap.reactions if snap else 0,
                'likes': snap.likes if snap else 0,
                'comments': snap.comments if snap else 0,
                'shares': snap.shares if snap else 0,
                'views': snap.views if snap else 0,
                'reach': snap.reach if snap else 0,
                'impressions': snap.impressions if snap else 0,
                'clicks': snap.clicks if snap else 0,
                'totalEngagement': snap.total_engagement if snap else 0,
                'engagementRate': snap.engagement_rate if snap else None,
            })
            
        return Response({
            'posts': result,
            'total': total,
            'page': page,
            'limit': limit,
            'totalPages': (total + limit - 1) // limit
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

