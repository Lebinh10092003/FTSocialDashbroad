from datetime import date, datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from examination.models import Candidate, CandidateParticipation, ExamSession, LogNote, RoundResult
from examination.sync import clean_txt, history_from_sheet_row, merged_headers, normalise_str, parse_dob, resolve_column_indices


def normalise_birth_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return parse_dob(value)


class Command(BaseCommand):
    help = 'Repair stored round scores from an official local XLSX without creating candidates.'

    def add_arguments(self, parser):
        parser.add_argument('--file', required=True, help='Path to the official XLSX source')
        parser.add_argument('--session', required=True, help='Target exam session id')
        parser.add_argument('--sheet', help='Workbook sheet name; defaults to the first sheet')
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        source = Path(options['file'])
        if not source.is_file():
            raise CommandError(f'Workbook not found: {source}')
        try:
            session = ExamSession.objects.get(id=options['session'])
        except ExamSession.DoesNotExist as error:
            raise CommandError('Exam session was not found.') from error

        workbook = load_workbook(source, read_only=True, data_only=True)
        if options.get('sheet'):
            if options['sheet'] not in workbook.sheetnames:
                raise CommandError(f"Sheet was not found: {options['sheet']}")
            sheet = workbook[options['sheet']]
        else:
            sheet = workbook[workbook.sheetnames[0]]
        grid = [list(row) for row in sheet.iter_rows(values_only=True)]
        header_index = next((index for index, row in enumerate(grid) if any(
            'hovaten' in normalise_str(cell)
            for cell in row
        )), None)
        if header_index is None:
            raise CommandError('Candidate header row was not found.')
        headers = merged_headers(grid, header_index)
        columns = resolve_column_indices(headers)
        profile_columns = {}
        profile_labels = {'code': 'mahoso', 'name': 'hovatenthisinh', 'dob': 'ngaysinh'}
        for field, label in profile_labels.items():
            profile_columns[field] = next((index for index, header in enumerate(headers)
                if label in normalise_str(str(header).split(':')[-1])), None)
        if any(index is None for index in profile_columns.values()):
            raise CommandError('Profile code, name, or birth-date columns were not found.')

        configured_rounds = [str(item.get('name') or '').strip() for item in (session.rounds or []) if isinstance(item, dict)]
        changed = unmatched = missing_participation = skipped = 0
        for row in grid[header_index + 1:]:
            if not row:
                continue
            def raw_cell(field):
                index = profile_columns.get(field, columns.get(field))
                return row[index] if index is not None and index < len(row) else ''

            def cell(field):
                return clean_txt(raw_cell(field))

            code, name, birth_date = cell('code'), cell('name'), normalise_birth_date(raw_cell('dob'))
            if not name:
                continue
            candidate = Candidate.objects.filter(code__iexact=code).first() if code else None
            if candidate is None and birth_date:
                candidates = Candidate.objects.filter(name__iexact=name, birth_date=birth_date)
                candidate = candidates.first() if candidates.count() == 1 else None
            if candidate is None and birth_date:
                # Source files can differ only in capitalization, spacing, or accents.
                # Keep this fallback strict by still requiring the full birth date.
                matches = [
                    item for item in Candidate.objects.filter(birth_date=birth_date)
                    if normalise_str(item.name) == normalise_str(name)
                ]
                candidate = matches[0] if len(matches) == 1 else None
            if candidate is None:
                # Some historical files have no FT code or birth date. It is still
                # safe to repair when exactly one same-name profile is already
                # enrolled in this exact session; never select an outside profile.
                same_name_in_session = [
                    item for item in Candidate.objects.all()
                    if normalise_str(item.name) == normalise_str(name)
                    and CandidateParticipation.objects.filter(candidate=item, session=session).exists()
                ]
                candidate = same_name_in_session[0] if len(same_name_in_session) == 1 else None
            if candidate is None:
                unmatched += 1
                continue
            participation = CandidateParticipation.objects.filter(candidate=candidate, session=session).first()
            if participation is None:
                missing_participation += 1
                continue

            results = list(participation.round_results.order_by('id'))
            for position, history in enumerate(history_from_sheet_row(headers, row)):
                score = clean_txt(history.get('score'))
                if not score:
                    skipped += 1
                    continue
                source_round = clean_txt(history.get('round'))
                result = RoundResult.objects.filter(participation=participation, round_name=source_round).first()
                if result is None and position < len(configured_rounds):
                    result = RoundResult.objects.filter(participation=participation, round_name=configured_rounds[position]).first()
                if result is None and position < len(results):
                    result = results[position]
                if result is None or result.score == score:
                    continue
                if not options['dry_run']:
                    previous = result.score or 'empty'
                    result.score = score
                    result.save(update_fields=['score', 'updated_at'])
                    LogNote.objects.create(
                        key=f'session-{session.id}:score-repair:{candidate.code}:{result.id}',
                        entity_key=f'candidate-{candidate.code}',
                        content=f'System repaired score for {result.round_name}: {previous} -> {score}.',
                        updated_by='FT Workspace system',
                        system=True,
                    )
                changed += 1

        mode = 'DRY RUN' if options['dry_run'] else 'UPDATED'
        self.stdout.write(self.style.SUCCESS(
            f'{mode}: scores={changed}, unmatched_candidates={unmatched}, '
            f'missing_participation={missing_participation}, no_score_rows={skipped}'
        ))