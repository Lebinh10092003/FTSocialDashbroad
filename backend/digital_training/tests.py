import io
import re
from collections import Counter
from datetime import date, datetime, time, timedelta
from zoneinfo import ZoneInfo
from zipfile import ZipFile
from unittest.mock import patch

from django.test import TestCase
from django.contrib.auth import get_user_model
from django.utils import timezone
from openpyxl import Workbook
from rest_framework.test import APIClient
from rest_framework.authtoken.models import Token

from authentication.models import Department, JobTitle, UserProfile
from .assessment_service import generate_variants_from_import, parse_assessment_workbook
from .completion_service import complete_past_training_schedules
from .models import TrainingAssessment, TrainingClass, TrainingCustomerMeeting, TrainingFinanceEntry, TrainingLead, TrainingPartner, TrainingProduct, TrainingProductSubscription, TrainingSession, TrainingSurvey
from .serializers import TrainingAssessmentSerializer, TrainingClassSerializer, TrainingCustomerMeetingSerializer, TrainingPartnerSerializer, TrainingProductSerializer, TrainingProductSubscriptionSerializer, TrainingSessionSerializer, TrainingSurveySerializer


class TrainingSurveySerializerTests(TestCase):
    def setUp(self):
        self.partner = TrainingPartner.objects.create(name="Khách hàng thử nghiệm")
        self.session = TrainingSession.objects.create(
            title="Buổi tập huấn thử nghiệm",
            partner=self.partner.name,
            partner_ref=self.partner,
            status="planned",
        )

    def test_creating_training_session_does_not_create_survey(self):
        TrainingSession.objects.create(title="Lịch mới", status="planned")

        self.assertEqual(TrainingSurvey.objects.count(), 0)

    def test_manual_survey_requires_training_session(self):
        serializer = TrainingSurveySerializer(
            data={
                "title": "Khảo sát cuối buổi",
                "form_type": "end_session",
                "notes": "",
            }
        )

        self.assertFalse(serializer.is_valid())
        self.assertIn("session", serializer.errors)

    def test_customer_is_derived_from_selected_training_session(self):
        serializer = TrainingSurveySerializer(
            data={
                "title": "Khảo sát cuối buổi",
                "form_type": "end_session",
                "session": self.session.pk,
                "notes": "",
            }
        )

        self.assertTrue(serializer.is_valid(), serializer.errors)
        survey = serializer.save()
        self.assertEqual(survey.session, self.session)
        self.assertEqual(survey.partner, self.partner)

class OtherWorkScheduleSerializerTests(TestCase):
    def test_other_schedule_keeps_its_selected_type(self):
        serializer = TrainingCustomerMeetingSerializer(
            data={
                "title": "Họp kế hoạch quý",
                "schedule_type": "other",
                "activity_type": "Họp nội bộ",
                "date": "2026-07-29",
                "status": "planned",
            }
        )

        self.assertTrue(serializer.is_valid(), serializer.errors)
        item = serializer.save()
        self.assertEqual(item.schedule_type, "other")
        self.assertEqual(item.activity_type, "Họp nội bộ")


class ScheduleCompletionTests(TestCase):
    def test_daily_completion_updates_only_schedules_that_have_ended(self):
        now = datetime(2026, 8, 1, 6, 0, tzinfo=ZoneInfo("Asia/Ho_Chi_Minh"))
        old_session = TrainingSession.objects.create(
            title="Buoi cu", session_date=date(2026, 7, 31), status="planned"
        )
        ended_today = TrainingSession.objects.create(
            title="Buoi som", session_date=date(2026, 8, 1), end_time=time(5, 30), status="planned"
        )
        future_today = TrainingSession.objects.create(
            title="Buoi chua ket thuc", session_date=date(2026, 8, 1), end_time=time(9, 0), status="planned"
        )
        cancelled = TrainingSession.objects.create(
            title="Buoi da huy", session_date=date(2026, 7, 31), status="cancelled"
        )
        old_meeting = TrainingCustomerMeeting.objects.create(
            title="Lich cong tac cu", meeting_date=date(2026, 7, 31), status="planned"
        )

        result = complete_past_training_schedules(now)

        self.assertEqual(result, {"sessions": 2, "customer_meetings": 1, "total": 3})
        old_session.refresh_from_db()
        ended_today.refresh_from_db()
        future_today.refresh_from_db()
        cancelled.refresh_from_db()
        old_meeting.refresh_from_db()
        self.assertEqual(old_session.status, "completed")
        self.assertEqual(ended_today.status, "completed")
        self.assertEqual(future_today.status, "planned")
        self.assertEqual(cancelled.status, "cancelled")
        self.assertEqual(old_meeting.status, "completed")
        self.assertEqual(complete_past_training_schedules(now)["total"], 0)

    def test_loading_sessions_repairs_stale_past_status_immediately(self):
        old_session = TrainingSession.objects.create(
            title="Lich hom qua",
            session_date=timezone.localdate() - timedelta(days=1),
            status="planned",
        )

        response = APIClient().get("/api/digital-training/sessions")

        self.assertEqual(response.status_code, 200)
        old_session.refresh_from_db()
        self.assertEqual(old_session.status, "completed")
class TrainingPartnerLocationTests(TestCase):
    def test_partner_keeps_province_and_ward_as_filterable_fields(self):
        serializer = TrainingPartnerSerializer(
            data={
                "name": "Trường thử nghiệm",
                "partner_type": "Khối Giáo dục",
                "partner_subtype": "THCS",
                "province": "Hà Nội",
                "ward": "Việt Hưng",
                "products": ["Tập huấn"],
            }
        )

        self.assertTrue(serializer.is_valid(), serializer.errors)
        partner = serializer.save()
        self.assertEqual(partner.province, "Hà Nội")
        self.assertEqual(partner.ward, "Việt Hưng")

    def test_training_contents_are_stored_on_each_class(self):
        partner = TrainingPartner.objects.create(name="Khách hàng nhiều lớp")
        serializer = TrainingClassSerializer(
            data={
                "partner": partner.pk,
                "name": "Lớp 1",
                "planned_sessions": 0,
                "training_contents": ["Dashboard", "Dashboard", "AI"],
            }
        )

        self.assertTrue(serializer.is_valid(), serializer.errors)
        training_class = serializer.save()
        self.assertEqual(training_class.training_contents, ["Dashboard", "AI"])


class TrainingProductManagementTests(TestCase):
    def setUp(self):
        self.partner = TrainingPartner.objects.create(
            name="Khach hang san pham",
            partner_type="Khoi Giao duc",
            partner_subtype="THPT",
        )

    def test_product_and_subscription_expose_quantity_and_remaining_status(self):
        product_serializer = TrainingProductSerializer(data={"name": "San pham moi", "product_type": "service", "description": "Mo ta"})
        self.assertTrue(product_serializer.is_valid(), product_serializer.errors)
        product = product_serializer.save()
        self.assertEqual(product.code, "san-pham-moi")
        self.assertEqual(product.product_type, "service")
        subscription_serializer = TrainingProductSubscriptionSerializer(data={
            "partner": self.partner.pk,
            "product": product.pk,
            "quantity": 25,
            "starts_at": timezone.localdate() - timedelta(days=30),
            "expires_at": timezone.localdate() + timedelta(days=14),
            "status": "active",
            "notes": "Gia han trong thang",
        })
        self.assertTrue(subscription_serializer.is_valid(), subscription_serializer.errors)
        subscription = subscription_serializer.save()
        data = TrainingProductSubscriptionSerializer(subscription).data
        self.assertEqual(data["effective_status"], "expiring")
        self.assertEqual(data["days_remaining"], 14)
        self.assertEqual(data["quantity"], 25)

    def test_product_subscription_is_unique_per_customer(self):
        product = TrainingProduct.objects.create(name="Unique product", code="unique-product")
        TrainingProductSubscription.objects.create(partner=self.partner, product=product, quantity=1)
        duplicate = TrainingProductSubscriptionSerializer(data={
            "partner": self.partner.pk,
            "product": product.pk,
            "quantity": 2,
        })
        self.assertFalse(duplicate.is_valid())
        self.assertIn("non_field_errors", duplicate.errors)

    def test_partner_legacy_products_create_product_subscriptions(self):
        ai_product, _ = TrainingProduct.objects.get_or_create(
            code="ai-dung-chung",
            defaults={"name": "AI dùng chung", "display_order": 2},
        )
        serializer = TrainingPartnerSerializer(data={
            "name": "Khach hang dong bo san pham",
            "products": ["AI có bản quyền dùng chung"],
            "ai_account_count": 18,
        })
        self.assertTrue(serializer.is_valid(), serializer.errors)
        partner = serializer.save()
        subscription = TrainingProductSubscription.objects.get(partner=partner, product=ai_product)
        self.assertEqual(subscription.quantity, 18)
        self.assertEqual(subscription.status, "active")

class TrainingSessionLocationTests(TestCase):
    def test_new_session_has_no_automatic_responsible_staff(self):
        serializer = TrainingSessionSerializer(
            data={
                "title": "Buổi chưa phân công",
                "date": "2026-08-05",
                "status": "planned",
            }
        )

        self.assertTrue(serializer.is_valid(), serializer.errors)
        session = serializer.save()
        self.assertEqual(session.staff_name, "")

    def test_partial_schedule_sync_does_not_overwrite_responsible_staff(self):
        session = TrainingSession.objects.create(
            title="Buổi 4 · UBP Giảng Võ · Lớp 2",
            staff_name="Ms Liên, Ms Phương",
            status="planned",
        )
        serializer = TrainingSessionSerializer(
            session,
            data={"location": "525 Kim Mã", "notes": "Đồng bộ lịch lớp."},
            partial=True,
        )

        self.assertTrue(serializer.is_valid(), serializer.errors)
        updated = serializer.save()
        self.assertEqual(updated.staff_name, "Ms Liên, Ms Phương")

    def test_location_is_stored_on_the_calendar_session(self):
        serializer = TrainingSessionSerializer(
            data={
                "title": "Buổi 1",
                "date": "2026-08-05",
                "start_time": "08:00",
                "end_time": "11:00",
                "location": "Hội trường A / https://meet.example.test",
                "status": "planned",
            }
        )

        self.assertTrue(serializer.is_valid(), serializer.errors)
        session = serializer.save()
        self.assertEqual(session.location, "Hội trường A / https://meet.example.test")
        self.assertEqual(TrainingSessionSerializer(session).data["location"], session.location)

    def test_shared_schedule_keeps_each_session_location(self):
        serializer = TrainingPartnerSerializer(
            data={
                "name": "Khách hàng có lịch theo địa điểm",
                "training_schedule": [
                    {
                        "date": "2026-08-05",
                        "start_time": "08:00",
                        "location": "Phòng 201",
                        "unscheduled": False,
                    },
                    {
                        "date": "2026-08-06",
                        "start_time": "13:30",
                        "location": "https://meet.example.test/buoi-2",
                        "unscheduled": False,
                    },
                ],
            }
        )

        self.assertTrue(serializer.is_valid(), serializer.errors)
        partner = serializer.save()
        self.assertEqual(partner.training_schedule[0]["location"], "Phòng 201")
        self.assertEqual(partner.training_schedule[1]["location"], "https://meet.example.test/buoi-2")


class TrainingAssessmentTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.partner = TrainingPartner.objects.create(name="UBP Giảng Võ")
        self.training_class = TrainingClass.objects.create(partner=self.partner, name="Lớp 2")
        self.assessment = TrainingAssessment.objects.create(
            title="Bài cuối học phần",
            partner=self.partner,
            training_class=self.training_class,
            status="published",
            duration_minutes=20,
            questions=[
                {
                    "id": f"q-{variant}",
                    "variant": variant,
                    "order": 1,
                    "type": "single_choice",
                    "text": "2 + 2 bằng bao nhiêu?",
                    "options": [{"key": "A", "text": "3"}, {"key": "B", "text": "4"}],
                    "correct_answers": ["B"],
                    "points": 1,
                    "required": True,
                }
                for variant in ["Đề 1", "Đề 2", "Đề 3", "Đề 4", "Đề 5"]
            ],
        )

    def test_public_url_uses_partner_and_class_slug(self):
        self.assertEqual(self.assessment.public_slug, "ubp-giang-vo-lop-2")

    def test_multiple_assessment_rounds_are_allowed_per_partner_class(self):
        serializer = TrainingAssessmentSerializer(data={
            "title": "Bài bị trùng",
            "training_class": self.training_class.pk,
            "duration_minutes": 15,
            "status": "draft",
            "questions": [],
        })
        self.assertTrue(serializer.is_valid(), serializer.errors)
        created = serializer.save()
        self.assertNotEqual(created.public_slug, self.assessment.public_slug)

    def test_variant_assignment_stays_balanced_behind_one_link(self):
        variants = []
        for index in range(13):
            response = self.client.post(
                f"/api/training-assessments/{self.assessment.public_slug}/start",
                {
                    "respondent_name": f"Người học {index}",
                    "email": f"learner{index}@example.test",
                },
                format="json",
            )
            self.assertEqual(response.status_code, 201, response.data)
            variants.append(response.data["variant"])
            self.assertNotIn("correct_answers", response.data["questions"][0])
        counts = Counter(variants)
        self.assertEqual(len(counts), 5)
        self.assertLessEqual(max(counts.values()) - min(counts.values()), 1)

    def test_submit_is_scored_on_server(self):
        start = self.client.post(
            f"/api/training-assessments/{self.assessment.public_slug}/start",
            {"respondent_name": "Nguyễn A", "email": "a@example.test"},
            format="json",
        )
        question_id = start.data["questions"][0]["id"]
        response = self.client.patch(
            f"/api/training-assessment-attempts/{start.data['access_token']}",
            {"answers": {question_id: "B"}, "submit": True},
            format="json",
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["status"], "submitted")
        self.assertEqual(float(response.data["score"]), 1)

    def test_server_rejects_submission_with_required_questions_unanswered(self):
        start = self.client.post(
            f"/api/training-assessments/{self.assessment.public_slug}/start",
            {"respondent_name": "Nguyễn A", "email": "missing@example.test"},
            format="json",
        )
        response = self.client.patch(
            f"/api/training-assessment-attempts/{start.data['access_token']}",
            {"submit": True},
            format="json",
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertIn("câu bắt buộc", response.data["error"])

    def test_required_matching_question_must_have_every_pair(self):
        self.assessment.questions = [{
            "id": "match-1",
            "variant": "Đề 1",
            "order": 1,
            "type": "matching",
            "text": "Ghép đủ hai cặp",
            "options": [
                {"key": "1", "text": "Trái 1", "match_text": "Phải A"},
                {"key": "2", "text": "Trái 2", "match_text": "Phải B"},
            ],
            "correct_answers": ["1-A", "2-B"],
            "points": 2,
            "required": True,
        }]
        self.assessment.save(update_fields=["questions", "updated_at"])
        start = self.client.post(
            f"/api/training-assessments/{self.assessment.public_slug}/start",
            {"respondent_name": "Nguyễn Ghép", "email": "matching@example.test"},
            format="json",
        )
        response = self.client.patch(
            f"/api/training-assessment-attempts/{start.data['access_token']}",
            {"answers": {"match-1": {"1": "A"}}, "submit": True},
            format="json",
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertIn("câu bắt buộc", response.data["error"])

    def test_expired_attempt_returns_timed_out_payload_instead_of_save_error(self):
        start = self.client.post(
            f"/api/training-assessments/{self.assessment.public_slug}/start",
            {"respondent_name": "Nguyễn Hết Giờ", "email": "expired@example.test"},
            format="json",
        )
        attempt = self.assessment.attempts.get(access_token=start.data["access_token"])
        attempt.expires_at = timezone.now() - timedelta(seconds=1)
        attempt.save(update_fields=["expires_at", "updated_at"])

        response = self.client.patch(
            f"/api/training-assessment-attempts/{start.data['access_token']}",
            {"answers": {}, "submit": True},
            format="json",
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["status"], "timed_out")

    def test_participant_list_assigns_fixed_variant_and_resumes_active_attempt(self):
        assigned_variant = self.assessment.questions[2]['variant']
        self.assessment.participants = [{
            "code": "GV-001",
            "name": "Teacher One",
            "email": "teacher@example.test",
            "phone": "",
            "organization": "School A",
            "group": "THPT",
            'variant': assigned_variant,
        }]
        self.assessment.save(update_fields=["participants", "updated_at"])
        first = self.client.post(
            f"/api/training-assessments/{self.assessment.public_slug}/start",
            {"participant_code": "GV-001"},
            format="json",
        )
        second = self.client.post(
            f"/api/training-assessments/{self.assessment.public_slug}/start",
            {"participant_code": "GV-001"},
            format="json",
        )

        self.assertEqual(first.status_code, 201, first.data)
        self.assertEqual(first.data['variant'], assigned_variant)
        self.assertEqual(second.status_code, 200, second.data)
        self.assertEqual(second.data["access_token"], first.data["access_token"])

    def test_open_assessment_resumes_across_browsers_only_for_exact_identity(self):
        first = self.client.post(
            f"/api/training-assessments/{self.assessment.public_slug}/start",
            {
                "respondent_name": "  Nguyễn   Văn A ",
                "email": "Learner@Example.Test",
                "phone": "090-123-4567",
            },
            format="json",
        )
        resumed = self.client.post(
            f"/api/training-assessments/{self.assessment.public_slug}/start",
            {
                "respondent_name": "nguyễn văn a",
                "email": "learner@example.test",
                "phone": "090 123 4567",
            },
            format="json",
        )

        self.assertEqual(first.status_code, 201, first.data)
        self.assertEqual(resumed.status_code, 200, resumed.data)
        self.assertTrue(resumed.data["resumed"])
        self.assertEqual(resumed.data["access_token"], first.data["access_token"])

        mismatch = self.client.post(
            f"/api/training-assessments/{self.assessment.public_slug}/start",
            {
                "respondent_name": "Tên Không Khớp",
                "email": "learner@example.test",
                "phone": "0901234567",
            },
            format="json",
        )
        self.assertEqual(mismatch.status_code, 400, mismatch.data)
        self.assertIn("chưa khớp", mismatch.data["error"])

    def test_reset_requires_exact_identity_and_clears_current_attempt(self):
        first = self.client.post(
            f"/api/training-assessments/{self.assessment.public_slug}/start",
            {
                "respondent_name": "Nguyễn Văn A",
                "email": "learner@example.test",
                "phone": "0901234567",
            },
            format="json",
        )
        question_id = first.data["questions"][0]["id"]
        self.client.patch(
            f"/api/training-assessment-attempts/{first.data['access_token']}",
            {"answers": {question_id: "B"}},
            format="json",
        )
        reset = self.client.post(
            f"/api/training-assessments/{self.assessment.public_slug}/start",
            {
                "respondent_name": "Nguyễn Văn A",
                "email": "learner@example.test",
                "phone": "090-123-4567",
                "reset": True,
            },
            format="json",
        )

        self.assertEqual(reset.status_code, 200, reset.data)
        self.assertTrue(reset.data["reset_performed"])
        self.assertEqual(reset.data["answers"], {})
        self.assertEqual(reset.data["access_token"], first.data["access_token"])

    def test_storage_config_rejects_unknown_folder_template_variables(self):
        serializer = TrainingAssessmentSerializer(
            instance=self.assessment,
            data={
                "storage_config": {
                    "participant_folder_template": "{respondent_name}-{secret}",
                },
            },
            partial=True,
        )

        self.assertFalse(serializer.is_valid())
        self.assertIn("storage_config", serializer.errors)

    def test_autosave_accepts_structured_answers_and_progress(self):
        start = self.client.post(
            f"/api/training-assessments/{self.assessment.public_slug}/start",
            {"respondent_name": "Learner", "email": "learner@example.test"},
            format="json",
        )
        question_id = start.data["questions"][0]["id"]

        response = self.client.patch(
            f"/api/training-assessment-attempts/{start.data['access_token']}",
            {
                "answers": {question_id: ["A", "B"]},
                "progress": {"current_question_id": question_id, "reviewed_question_ids": [question_id]},
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["answers"][question_id], ["A", "B"])
        self.assertEqual(response.data["progress"]["current_question_id"], question_id)
        self.assertEqual(response.data["progress"]["reviewed_question_ids"], [question_id])

    def test_xlsx_parser_supports_question_bank_schema_and_interactions(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "THPT"
        sheet.append([
            "order", "topic", "knowledge type", "question type", "difficulty",
            "question", "media url", "option 1", "option 2", "option 3",
            "option 4", "option 5", "answer", "answer image", "points",
        ])
        sheet.append([1, "Topic A", "Theory", "multiple choice", "Medium", "Choose values", "", "One", "Two", "Three", "", "", "1;3", "", 2])
        sheet.append([2, "Topic B", "Theory", "matching", "Hard", "Match values", "", "Left 1 | Right 1", "Left 2 | Right 2", "", "", "", "1-B;2-A", "", 2])
        sheet.append([3, "Topic C", "Theory", "ordering", "Hard", "Order values", "", "Step 1", "Step 2", "Step 3", "", "", "2-1-3", "", 2])
        sheet.append([4, "Practice", "Practice", "link upload", "Hard", "Submit product", "", "", "", "", "", "", "Manual", "", 6])
        buffer = io.BytesIO()
        workbook.save(buffer)

        result = parse_assessment_workbook(buffer.getvalue(), "bank.xlsx")

        self.assertEqual(result["errors"], [])
        self.assertEqual([item["type"] for item in result["questions"]], [
            "multiple_choice", "matching", "ordering", "practical_submission",
        ])
        self.assertEqual(result["questions"][0]["correct_answers"], ["1", "3"])
        self.assertEqual(result["questions"][1]["options"][0]["text"], "Left 1")
        self.assertEqual(result["questions"][1]["options"][0]["match_text"], "Right 1")
        self.assertEqual(result["questions"][0]["audience_group"], "THPT")
        self.assertTrue(all(item["question_code"] for item in result["questions"]))

    def test_xlsx_parser_supports_variant_column(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Nguồn câu hỏi"
        sheet.append(["Mã đề", "STT", "Loại câu", "Câu hỏi", "A", "B", "Đáp án", "Điểm"])
        sheet.append(["Đề 1", 1, "Trắc nghiệm", "Câu một", "Sai", "Đúng", "B", 1])
        sheet.append(["Đề 2", 1, "Trả lời ngắn", "Câu hai", "", "", "Hà Nội", 2])
        buffer = io.BytesIO()
        workbook.save(buffer)
        result = parse_assessment_workbook(buffer.getvalue(), "questions.xlsx")
        self.assertEqual(result["question_count"], 2)
        self.assertEqual([item["name"] for item in result["variants"]], ["Đề 1", "Đề 2"])
        self.assertEqual(result["errors"], [])

    def test_xlsx_parser_supports_missing_worksheet_dimension_metadata(self):
        workbook = Workbook()
        guide = workbook.active
        guide.title = 'Guide'
        guide.append(['GUIDE'])
        sheet = workbook.create_sheet('De 1')
        sheet.append(['order', 'type', 'question', 'A', 'B', 'answer', 'points'])
        sheet.append([1, 'single_choice', 'Valid question', 'Wrong', 'Right', 'B', 1])
        source = io.BytesIO()
        workbook.save(source)

        dimensionless = io.BytesIO()
        with ZipFile(io.BytesIO(source.getvalue())) as archive, ZipFile(dimensionless, 'w') as output:
            for entry in archive.infolist():
                content = archive.read(entry.filename)
                if entry.filename.startswith('xl/worksheets/sheet'):
                    content = re.sub(rb'<dimension\b[^>]*/>', b'', content, count=1)
                output.writestr(entry, content)

        result = parse_assessment_workbook(dimensionless.getvalue(), 'dimensionless.xlsx')

        self.assertEqual(result['question_count'], 1)
        self.assertEqual(len(result['variants']), 1)
        self.assertEqual(result['errors'], [])

    def test_import_generation_balances_usage_and_answer_keys(self):
        source_questions = [
            {
                "id": f"bank-{index}",
                "variant": "Đề 1",
                "order": index,
                "type": "single_choice",
                "text": f"Câu hỏi nguồn {index}",
                "options": [
                    {"key": "A", "text": "Đúng"},
                    {"key": "B", "text": "Sai 1"},
                    {"key": "C", "text": "Sai 2"},
                    {"key": "D", "text": "Sai 3"},
                ],
                "correct_answers": ["A"],
                "points": 1,
                "required": True,
                "category": "Kiến thức",
                "difficulty": "Trung bình",
            }
            for index in range(1, 13)
        ]
        result = generate_variants_from_import(source_questions, 4, 5, seed=20260730)
        self.assertEqual(len(result["variants"]), 4)
        self.assertEqual(result["question_count"], 20)
        source_usage = Counter(item["source_question_id"] for item in result["questions"])
        self.assertLessEqual(max(source_usage.values()) - min(source_usage.values()), 1)
        for variant in result["variants"]:
            questions = [
                item for item in result["questions"]
                if item["variant"] == variant["name"]
            ]
            self.assertEqual(len({item["source_question_id"] for item in questions}), 5)
            answer_counts = Counter(item["correct_answers"][0] for item in questions)
            self.assertLessEqual(max(answer_counts.values()) - min(answer_counts.values()), 1)

    def test_import_generation_removes_duplicate_questions(self):
        source_questions = [
            {
                "id": f"bank-{index}",
                "type": "short_answer",
                "text": "Câu bị trùng" if index < 3 else f"Câu {index}",
                "options": [],
                "correct_answers": ["Đúng"],
                "points": 1,
                "required": True,
            }
            for index in range(1, 6)
        ]
        result = generate_variants_from_import(source_questions, 2, 3, seed=1)
        self.assertEqual(result["source_question_count"], 4)
        self.assertTrue(any("Đã bỏ 1 câu trùng" in warning for warning in result["warnings"]))

    def test_import_generation_honors_topic_and_difficulty_structure(self):
        source_questions = []
        for category, difficulty, count in [("Topic A", "Easy", 5), ("Topic B", "Hard", 5)]:
            for index in range(count):
                source_questions.append({
                    "id": f"{category}-{difficulty}-{index}",
                    "type": "short_answer",
                    "text": f"{category} {difficulty} question {index}",
                    "options": [],
                    "correct_answers": ["ok"],
                    "points": 1,
                    "required": True,
                    "category": category,
                    "difficulty": difficulty,
                })
        structure = [
            {"category": "Topic A", "difficulty": "Easy", "count": 2},
            {"category": "Topic B", "difficulty": "Hard", "count": 3},
        ]
        result = generate_variants_from_import(source_questions, 3, 5, seed=7, structure=structure)
        for variant in result["variants"]:
            questions = [item for item in result["questions"] if item["variant"] == variant["name"]]
            counts = Counter((item["category"], item["difficulty"]) for item in questions)
            self.assertEqual(counts[("Topic A", "Easy")], 2)
            self.assertEqual(counts[("Topic B", "Hard")], 3)
        self.assertEqual(sum(item["count"] for item in result["generation_config"]["structure"]), 5)

    def test_import_generation_honors_knowledge_type_and_question_type_structure(self):
        source_questions = []
        for knowledge_type, question_type, difficulty in [
            ("Theory", "short_answer", "Easy"),
            ("Practice", "practical_submission", "Hard"),
        ]:
            for index in range(1, 5):
                source_questions.append({
                    "id": f"{knowledge_type}-{question_type}-{index}",
                    "type": question_type,
                    "text": f"{knowledge_type} {question_type} question {index}",
                    "options": [],
                    "correct_answers": ["ok"],
                    "points": 1,
                    "required": True,
                    "knowledge_type": knowledge_type,
                    "difficulty": difficulty,
                })
        structure = [
            {"knowledge_type": "Theory", "type": "short_answer", "difficulty": "Easy", "count": 2},
            {"knowledge_type": "Practice", "type": "practical_submission", "difficulty": "Hard", "count": 2},
        ]

        result = generate_variants_from_import(source_questions, 3, 4, seed=7, structure=structure)

        for variant in result["variants"]:
            questions = [item for item in result["questions"] if item["variant"] == variant["name"]]
            counts = Counter((item["knowledge_type"], item["type"], item["difficulty"]) for item in questions)
            self.assertEqual(counts[("Theory", "short_answer", "Easy")], 2)
            self.assertEqual(counts[("Practice", "practical_submission", "Hard")], 2)
    def test_xlsx_parser_uses_sheet_names_for_prepared_variants(self):
        workbook = Workbook()
        first = workbook.active
        first.title = "Đề 1"
        second = workbook.create_sheet("Đề 2")
        for sheet, question in [(first, "Câu của đề một"), (second, "Câu của đề hai")]:
            sheet.append(["STT", "Loại câu", "Câu hỏi", "A", "B", "Đáp án", "Điểm"])
            sheet.append([1, "Trắc nghiệm", question, "Sai", "Đúng", "B", 1])
        buffer = io.BytesIO()
        workbook.save(buffer)
        result = parse_assessment_workbook(buffer.getvalue(), "prepared.xlsx")
        self.assertEqual([item["name"] for item in result["variants"]], ["Đề 1", "Đề 2"])
        self.assertEqual(result["question_count"], 2)
        self.assertEqual(result["errors"], [])


class TrainingAssessmentImportPreviewTests(TestCase):
    def setUp(self):
        self.manager = UserProfile.objects.create(
            email="assessment-manager@example.test",
            name="Assessment Manager",
            role="MANAGER",
            access_modules=["digital-training"],
        )
        self.client = APIClient()
        self.client.force_authenticate(self.manager)

    def test_explicit_variant_count_overrides_participant_based_suggestion(self):
        source_questions = [{
            "id": "bank-1",
            "type": "short_answer",
            "text": "Question one",
            "options": [],
            "correct_answers": ["ok"],
            "points": 1,
            "required": True,
        }]
        parsed = {"questions": source_questions, "errors": []}
        with patch("digital_training.assessment_views.fetch_google_sheet", return_value=b"workbook"), patch(
            "digital_training.assessment_views.parse_assessment_workbook", return_value=parsed
        ):
            response = self.client.post(
                "/api/digital-training/assessments/import-preview",
                {
                    "google_sheet_url": "https://docs.google.com/spreadsheets/d/example",
                    "import_mode": "auto_generate",
                    "participant_count": 40,
                    "max_people_per_variant": 8,
                    "variant_count": 3,
                    "questions_per_variant": 1,
                    "seed": 1,
                },
                format="json",
            )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(len(response.data["variants"]), 3)
        self.assertEqual(response.data["generation_config"]["variant_count"], 3)

class TrainingFinancePermissionTests(TestCase):
    def client_for(self, email, role="EMPLOYEE", title_name="", department_name=""):
        title = JobTitle.objects.create(name=title_name) if title_name else None
        department = Department.objects.create(name=department_name) if department_name else None
        profile = UserProfile.objects.create(email=email, name=email, role=role, job_title=title, department=department)
        if department:
            profile.departments.add(department)
        django_user = get_user_model().objects.create_user(
            username=email,
            email=email,
            password="StrongPassword9921",
        )
        token = Token.objects.create(user=django_user).key
        client = APIClient()
        client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")
        return client

    def payload(self):
        return {
            "transaction_date": "2026-08-01",
            "entry_type": "income",
            "category": "Hop dong",
            "description": "Thu dot 1",
            "amount": 15000000,
            "status": "completed",
        }

    def test_accountant_and_admin_can_create_finance_entries(self):
        accountant = self.client_for("accountant@example.com", department_name="Phong Ke toan")
        response = accountant.post("/api/digital-training/finance-entries", self.payload(), format="json")
        self.assertEqual(response.status_code, 201, response.data)

        admin = self.client_for("finance-admin@example.com", role="ADMIN")
        response = admin.post("/api/digital-training/finance-entries", {
            **self.payload(),
            "entry_type": "expense",
            "description": "Chi van hanh",
        }, format="json")
        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(TrainingFinanceEntry.objects.count(), 2)

    def test_manager_and_director_can_view_but_cannot_edit(self):
        TrainingFinanceEntry.objects.create(**self.payload())
        manager = self.client_for("manager-finance@example.com", role="MANAGER")
        self.assertEqual(manager.get("/api/digital-training/finance-entries").status_code, 200)
        self.assertEqual(manager.post("/api/digital-training/finance-entries", self.payload(), format="json").status_code, 403)

        director = self.client_for("director@example.com", title_name="Giam doc")
        self.assertEqual(director.get("/api/digital-training/finance-entries").status_code, 200)
        self.assertEqual(director.patch("/api/digital-training/finance-entries/1", {"notes": "No edit"}, format="json").status_code, 403)

    def test_ordinary_employee_cannot_view_finance_entries(self):
        employee = self.client_for("employee-finance@example.com", title_name="Nhan vien ky thuat")
        self.assertEqual(employee.get("/api/digital-training/finance-entries").status_code, 403)

class TrainingLeadConversionTests(TestCase):
    def setUp(self):
        self.manager = UserProfile.objects.create(
            email="manager@fermat.vn", name="Quản lý", role="MANAGER", access_modules=["digital-training"]
        )
        self.client = APIClient()
        self.client.force_authenticate(self.manager)

    def test_guest_cannot_list_prospective_customers(self):
        response = APIClient().get("/api/digital-training/leads")
        self.assertEqual(response.status_code, 401)

    def test_lead_keeps_linked_meetings_after_conversion(self):
        lead = TrainingLead.objects.create(
            name="Trường THCS Mới", lead_type="Khối Giáo dục", representative="Cô Lan",
            phone="0900000000", email="lan@example.com", stage="negotiation",
        )
        meeting = TrainingCustomerMeeting.objects.create(
            title="Thương thảo hợp đồng", lead=lead, meeting_date=date(2026, 8, 5), status="planned"
        )
        response = self.client.post(f"/api/digital-training/leads/{lead.id}/convert", {}, format="json")
        self.assertEqual(response.status_code, 201, response.data)
        lead.refresh_from_db()
        meeting.refresh_from_db()
        self.assertEqual(lead.stage, "converted")
        self.assertEqual(lead.converted_partner.name, "Trường THCS Mới")
        self.assertEqual(meeting.lead_id, lead.id)
        self.assertEqual(response.data["partner"]["contact_person"], "Cô Lan")

    def test_conversion_rejects_duplicate_current_customer(self):
        TrainingPartner.objects.create(name="Existing school")
        lead = TrainingLead.objects.create(name="Existing school")

        response = self.client.post(f"/api/digital-training/leads/{lead.id}/convert", {}, format="json")

        self.assertEqual(response.status_code, 409, response.data)
        lead.refresh_from_db()
        self.assertIsNone(lead.converted_partner)

    def test_meeting_serializer_returns_linked_lead_name(self):
        lead = TrainingLead.objects.create(name="Đầu mối mới")
        meeting = TrainingCustomerMeeting.objects.create(
            title="Lịch gặp", lead=lead, meeting_date=date(2026, 8, 5), status="planned"
        )
        data = TrainingCustomerMeetingSerializer(meeting).data
        self.assertEqual(data["lead"], lead.id)
        self.assertEqual(data["lead_name"], lead.name)