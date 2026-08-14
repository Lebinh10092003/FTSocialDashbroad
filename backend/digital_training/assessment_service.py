import copy
import hashlib
import io
import json
import os
import random
import re
import secrets
import unicodedata
from collections import Counter
from decimal import Decimal

import requests
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from openpyxl import load_workbook

from integrations.google_sheets import build_sheets_service, extract_spreadsheet_id


def _key(value):
    text = str(value or "").strip().lower().replace("đ", "d")
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"[^a-z0-9]+", "", text)


COLUMN_ALIASES = {
    "question_code": {"macauhoi", "questioncode", "questionid", "idcauhoi"},
    "variant": {"made", "de", "variant", "version", "phienban"},
    "order": {"stt", "socau", "cau", "questionnumber", "order"},
    "knowledge_type": {"loaicauhoi", "lythuyetthuchanh", "knowledgetype"},
    "type": {"loaicau", "kieucauhoi", "dangcau", "type", "questiontype"},
    "text": {"cauhoi", "noidungcauhoi", "noidung", "question", "text"},
    "correct": {"dapan", "dapandung", "correctanswer", "answer"},
    "points": {"diem", "sodiem", "points", "score"},
    "required": {"batbuoc", "required"},
    "explanation": {"giaithich", "huongdan", "explanation"},
    "media_url": {"hinhanh", "anh", "image", "imageurl", "anhvideominhhoa", "anhvideominhhoaneuco", "media", "mediaurl"},
    "answer_image_url": {"anhdapan", "anhdapanneuco", "answerimage", "answerimageurl"},
    "category": {"chude", "nhomcau", "nhom", "category", "topic"},
    "difficulty": {"dokho", "mucdo", "difficulty", "level"},
}
for position, letter in enumerate("ABCDE", start=1):
    COLUMN_ALIASES[f"option_{position}"] = {
        str(position),
        letter.lower(),
        f"phuongan{position}",
        f"phuongan{letter.lower()}",
        f"dapan{letter.lower()}",
        f"option{letter.lower()}",
        f"option{position}",
    }


def _column_name(value):
    normalized = _key(value)
    for canonical, aliases in COLUMN_ALIASES.items():
        if normalized in aliases:
            return canonical
    return ""


def _question_type(value):
    normalized = _key(value)
    if normalized in {"tracnghiem", "tracnghiemmotdapan", "singlechoice", "chonmot", "mcq", ""}:
        return "single_choice"
    if normalized in {"tracnghiemnhieudapan", "multiplechoice", "chonnhieu", "multiselect"}:
        return "multiple_choice"
    if normalized in {"traloingan", "shortanswer", "shorttext", "tuluanngan", "diendapan", "diennoidung"}:
        return "short_answer"
    if normalized in {"ghepnoi", "matching", "match"}:
        return "matching"
    if normalized in {"sapxepthutu", "ordering", "sorting", "reorder"}:
        return "ordering"
    if normalized in {"fileupload", "uploadtep", "noptep", "tailen"}:
        return "file_upload"
    if normalized in {"taianh", "uploadanh", "upload", "thuchanh", "ganlinktaianh", "diendapanganlink", "ganlink", "linkupload"}:
        return "practical_submission"
    return normalized


def _is_practical_question(question):
    return (
        _key(question.get("knowledge_type")) in {"thuchanh", "practice", "practical"}
        or _question_type(question.get("type")) in {"practical_submission", "file_upload"}
    )


def _bool(value, default=True):
    if value in (None, ""):
        return default
    return _key(value) not in {"0", "false", "no", "khong", "khongbatbuoc"}


def _number(value, default=1):
    try:
        return max(0, float(value))
    except (TypeError, ValueError):
        return default


def _drive_reference(value):
    source = str(value or "").strip()
    if not source:
        return "", ""
    match = re.search(r"(?:/d/|[?&]id=)([a-zA-Z0-9_-]+)", source)
    if match:
        return match.group(1), source
    if re.fullmatch(r"[a-zA-Z0-9_-]{20,}", source):
        return source, f"https://drive.google.com/open?id={source}"
    return "", source


def _option_answer_key(value, option_keys):
    part = str(value or "").strip().upper()
    if part in option_keys:
        return part
    if part in "ABCDE":
        numeric = str(ord(part) - ord("A") + 1)
        return numeric if numeric in option_keys else ""
    return ""


def _split_matching_column(value, label_pattern):
    source = str(value or "").strip()
    matches = list(re.finditer(label_pattern, source, flags=re.IGNORECASE))
    if not matches:
        return []
    result = []
    for index, match in enumerate(matches):
        label = match.group(1).upper()
        text = source[match.end():matches[index + 1].start() if index + 1 < len(matches) else len(source)].strip()
        if text:
            result.append((label, text))
    return result


def _normalized_matching_options(options):
    """Support banks that place all left/right values in separate CỘT A/B cells."""
    original = [dict(option) for option in options if isinstance(option, dict)]
    if not original or all(str(option.get("match_text") or "").strip() for option in original):
        return original
    left_source = next((str(option.get("text") or "") for option in original if _key(option.get("text")).startswith("cota")), "")
    right_source = next((str(option.get("text") or "") for option in original if _key(option.get("text")).startswith("cotb")), "")
    left_items = _split_matching_column(left_source, r"(?:^|\s)([A-E])\s*[.)]")
    right_items = _split_matching_column(right_source, r"(?:^|\s)(\d{1,2})\s*[.)]")
    if len(left_items) < 2 or len(left_items) != len(right_items):
        return original
    return [
        {"key": str(index), "text": left_text, "match_text": right_text}
        for index, ((_, left_text), (_, right_text)) in enumerate(zip(left_items, right_items), start=1)
    ]


def _header_row(sheet):
    # Some valid XLSX generators omit worksheet ``<dimension>`` metadata.
    # In read-only mode openpyxl then leaves max_row/max_column as None until
    # the worksheet is scanned explicitly.
    if sheet.max_row is None:
        sheet.calculate_dimension(force=True)
    for row_number in range(1, min(sheet.max_row, 12) + 1):
        columns = {_column_name(cell.value) for cell in sheet[row_number]}
        supporting_columns = {"question_code", "order", "type", "correct", "points", "option_1"}
        if "text" in columns and columns.intersection(supporting_columns):
            return row_number
    return None


def parse_assessment_workbook(content, source_name=""):
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    questions = []
    errors = []
    warnings = []
    usable_sheets = []
    for sheet in workbook.worksheets:
        header_row = _header_row(sheet)
        if header_row:
            usable_sheets.append((sheet, header_row))
    for sheet, header_row in usable_sheets:
        columns = {}
        for index, cell in enumerate(sheet[header_row], start=1):
            canonical = _column_name(cell.value)
            if canonical and canonical not in columns:
                columns[canonical] = index
        # Read-only worksheets are forward-only.  Calling ``sheet.cell`` for
        # every field restarts the stream repeatedly, which makes a modest
        # Google workbook take minutes to parse.  Consume each source row once.
        for row_number, row_values in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            def value(name):
                column = columns.get(name)
                return row_values[column - 1] if column and column <= len(row_values) else None

            text = str(value("text") or "").strip()
            if not text:
                continue
            variant = str(value("variant") or (sheet.title if len(usable_sheets) > 1 else "Đề 1")).strip()
            question_type = _question_type(value("type"))
            question_code = str(value("question_code") or "").strip() or f"{_key(sheet.title)[:12]}-{row_number:04d}"
            if question_type not in {"single_choice", "multiple_choice", "short_answer", "matching", "ordering", "practical_submission", "file_upload"}:
                errors.append(f"{sheet.title}!{row_number}: loại câu “{value('type')}” chưa được hỗ trợ.")
                continue
            options = []
            for position in range(1, 6):
                option_text = str(value(f"option_{position}") or "").strip()
                if option_text:
                    options.append({"key": str(position), "text": option_text})
            if question_type == "matching":
                for option in options:
                    pair = re.split(r"\s*(?:=>|→|\|)\s*", option["text"], maxsplit=1)
                    if len(pair) == 2 and all(pair):
                        option["text"], option["match_text"] = pair
                options = _normalized_matching_options(options)
                if options and not all(option.get("match_text") for option in options):
                    warnings.append(
                        f"{sheet.title}!{row_number}: nên nhập mỗi cặp ghép theo dạng “vế trái | vế phải” để người làm thấy đủ nội dung."
                    )
            correct_raw = str(value("correct") or "").strip()
            correct = []
            if question_type in {"single_choice", "multiple_choice"}:
                if len(options) < 2:
                    errors.append(f"{sheet.title}!{row_number}: câu trắc nghiệm cần ít nhất 2 phương án.")
                raw_parts = [part.strip() for part in re.split(r"[,;|]", correct_raw) if part.strip()]
                option_keys = {option["key"] for option in options}
                for part in raw_parts:
                    matched_key = _option_answer_key(part, option_keys)
                    if matched_key:
                        correct.append(matched_key)
                        continue
                    matched = next((option["key"] for option in options if _key(option["text"]) == _key(part)), None)
                    if matched:
                        correct.append(matched)
                correct = list(dict.fromkeys(correct))
                if question_type == "single_choice" and len(correct) != 1:
                    errors.append(f"{sheet.title}!{row_number}: cần đúng 1 đáp án A–E.")
                elif question_type == "multiple_choice" and not correct:
                    errors.append(f"{sheet.title}!{row_number}: cần có ít nhất 1 đáp án đúng.")
            elif question_type in {"short_answer", "matching", "ordering"} and correct_raw:
                separator = r"[|;]" if question_type == "short_answer" else r"[|]"
                correct = [part.strip() for part in re.split(separator, correct_raw) if part.strip()]
            media_file_id, media_url = _drive_reference(value("media_url"))
            answer_image_file_id, answer_image_url = _drive_reference(value("answer_image_url"))
            question = {
                "id": f"{_key(variant) or 'de'}-{_key(question_code) or row_number}",
                "question_code": question_code,
                "variant": variant,
                "order": int(_number(value("order"), len(questions) + 1)),
                "type": question_type,
                "text": text,
                "options": options,
                "correct_answers": correct,
                "points": _number(value("points"), 1),
                "required": _bool(value("required"), True),
                "explanation": str(value("explanation") or "").strip(),
                "image_url": media_url,
                "media_url": media_url,
                "media_file_id": media_file_id,
                "answer_image_url": answer_image_url,
                "answer_image_file_id": answer_image_file_id,
                "knowledge_type": str(value("knowledge_type") or "").strip(),
                "category": str(value("category") or "").strip(),
                "difficulty": str(value("difficulty") or "").strip(),
                "audience_group": sheet.title,
                "source": f"{sheet.title}!{row_number}",
            }
            questions.append(question)
    if not usable_sheets:
        errors.append("Không tìm thấy cột “Câu hỏi” trong workbook.")
    variants = sorted({question["variant"] for question in questions}, key=str.casefold)
    counts = Counter(question["variant"] for question in questions)
    point_totals = {}
    for variant in variants:
        variant_questions = [question for question in questions if question["variant"] == variant]
        fingerprints = [_question_fingerprint(question) for question in variant_questions]
        if len(fingerprints) != len(set(fingerprints)):
            errors.append(f"{variant}: có câu hỏi bị trùng trong cùng một mã đề.")
        point_totals[variant] = sum(
            Decimal(str(question.get("points") or 0))
            for question in variant_questions
        )
    if len(variants) == 1:
        warnings.append("Workbook hiện có 1 mã đề. Có thể thêm cột “Mã đề” hoặc dùng mỗi sheet làm một đề.")
    if variants and len(set(counts.values())) > 1:
        warnings.append("Số câu giữa các mã đề chưa bằng nhau; hệ thống vẫn chia đề nhưng nên kiểm tra lại.")
    if variants and len(set(point_totals.values())) > 1:
        warnings.append("Tổng điểm giữa các mã đề chưa bằng nhau; nên điều chỉnh trước khi phát hành.")
    return {
        "source_name": source_name,
        "questions": questions,
        "variants": [{"name": name, "question_count": counts[name]} for name in variants],
        "question_count": len(questions),
        "errors": errors,
        "warnings": warnings,
    }


def _question_fingerprint(question):
    option_text = "|".join(_key(item.get("text")) for item in question.get("options") or [])
    return f"{_key(question.get('text'))}|{option_text}"


def _shuffle_options(question, rng, target_index):
    item = copy.deepcopy(question)
    if item.get("type") != "single_choice" or len(item.get("correct_answers") or []) != 1:
        return item
    options = item.get("options") or []
    correct_key = item["correct_answers"][0]
    correct_option = next((option for option in options if option.get("key") == correct_key), None)
    if not correct_option:
        return item
    distractors = [option for option in options if option is not correct_option]
    rng.shuffle(distractors)
    target_index = target_index % len(options)
    arranged = distractors[:]
    arranged.insert(target_index, correct_option)
    numeric_keys = all(str(option.get("key") or "").isdigit() for option in options)
    item["options"] = [
        {"key": str(index + 1) if numeric_keys else chr(65 + index), "text": option.get("text", "")}
        for index, option in enumerate(arranged)
    ]
    item["correct_answers"] = [str(target_index + 1) if numeric_keys else chr(65 + target_index)]
    return item


def generate_variants_from_import(questions, variant_count=5, questions_per_variant=20, seed=None, structure=None, topic_config=None, knowledge_config=None, score_config=None, difficulty_config=None):
    try:
        variant_count = int(variant_count)
        questions_per_variant = int(questions_per_variant)
    except (TypeError, ValueError):
        raise ValueError("Số mã đề và số câu mỗi đề phải là số nguyên.")
    if variant_count < 1 or variant_count > 200:
        raise ValueError("Số mã đề phải từ 1 đến 200.")
    if questions_per_variant < 1 or questions_per_variant > 200:
        raise ValueError("Số câu mỗi đề phải từ 1 đến 200.")

    unique_questions = []
    seen = set()
    duplicate_count = 0
    for question in questions:
        fingerprint = _question_fingerprint(question)
        if fingerprint in seen:
            duplicate_count += 1
            continue
        seen.add(fingerprint)
        unique_questions.append(question)
    if len(unique_questions) < questions_per_variant:
        raise ValueError(
            f"File nguồn chỉ có {len(unique_questions)} câu không trùng; "
            f"không đủ để tạo đề {questions_per_variant} câu."
        )

    def normalize_knowledge(value):
        normalized = str(value or "").strip().casefold()
        if normalized in {"lý thuyết", "ly thuyet", "theory"}:
            return "theory"
        if normalized in {"thực hành", "thuc hanh", "practice"}:
            return "practice"
        return normalized

    def normalize_difficulty(value):
        normalized = str(value or "").strip().casefold()
        if normalized in {"dễ", "de", "easy"}:
            return "easy"
        if normalized in {"trung bình", "trung binh", "medium"}:
            return "medium"
        if normalized in {"khó", "kho", "hard"}:
            return "hard"
        return normalized

    if knowledge_config and not isinstance(knowledge_config, dict):
        raise ValueError("Cơ cấu Lý thuyết/Thực hành không hợp lệ.")
    knowledge_rule = None
    if knowledge_config:
        try:
            theory_total = int(knowledge_config.get("theory") or 0)
            practice_total = int(knowledge_config.get("practice") or 0)
        except (TypeError, ValueError):
            raise ValueError("Số câu Lý thuyết/Thực hành phải là số nguyên.")
        if min(theory_total, practice_total) < 0 or theory_total + practice_total != questions_per_variant:
            raise ValueError("Tổng Lý thuyết + Thực hành phải bằng số câu mỗi đề.")
        knowledge_rule = {"theory": theory_total, "practice": practice_total}

    if score_config and not isinstance(score_config, dict):
        raise ValueError("Cấu hình điểm không hợp lệ.")
    score_rule = {}
    for knowledge in ("theory", "practice"):
        if score_config and score_config.get(knowledge) not in (None, ""):
            try:
                points = float(score_config[knowledge])
            except (TypeError, ValueError):
                raise ValueError("Điểm mỗi câu phải là một số.")
            if points < 0 or points > 1000:
                raise ValueError("Điểm mỗi câu phải từ 0 đến 1000.")
            score_rule[knowledge] = points

    if difficulty_config and not isinstance(difficulty_config, dict):
        raise ValueError("Cau hinh Do de khong hop le.")
    difficulty_rule = None
    if difficulty_config:
        try:
            easy_total = int(difficulty_config.get("easy") or 0)
            medium_total = int(difficulty_config.get("medium") or 0)
            hard_total = int(difficulty_config.get("hard") or 0)
        except (TypeError, ValueError):
            raise ValueError("So cau De/Trung binh/Kho phai la so nguyen.")
        if min(easy_total, medium_total, hard_total) < 0 or easy_total + medium_total + hard_total != questions_per_variant:
            raise ValueError("Tong De + Trung binh + Kho phai bang so cau moi de.")
        difficulty_rule = {"easy": easy_total, "medium": medium_total, "hard": hard_total}

    topic_rules = []
    seen_topics = set()
    for config in topic_config or []:
        if not isinstance(config, dict):
            raise ValueError("Cấu hình chủ đề không hợp lệ.")
        try:
            total = int(config.get("total") or 0)
        except (TypeError, ValueError):
            raise ValueError("Số câu trong cơ cấu chủ đề phải là số nguyên.")
        if total <= 0:
            continue
        category = str(config.get("category") or "").strip().casefold()
        if not category:
            raise ValueError("Vui lòng chọn một chủ đề hợp lệ.")
        if category in seen_topics:
            raise ValueError("Một chủ đề chỉ được cấu hình một lần.")
        has_knowledge = any(key in config for key in ("theory", "practice"))
        has_difficulty = any(key in config for key in ("easy", "medium", "hard"))
        try:
            theory = int(config.get("theory") or 0)
            practice = int(config.get("practice") or 0)
            easy = int(config.get("easy") or 0)
            medium = int(config.get("medium") or 0)
            hard = int(config.get("hard") or 0)
        except (TypeError, ValueError):
            raise ValueError("Số câu trong cơ cấu chủ đề phải là số nguyên.")
        if min(theory, practice, easy, medium, hard) < 0:
            raise ValueError("Số câu trong cơ cấu chủ đề không được âm.")
        if has_knowledge and theory + practice != total:
            raise ValueError("Tổng Lý thuyết + Thực hành phải bằng số câu của từng chủ đề.")
        if has_difficulty and easy + medium + hard != total:
            raise ValueError("Tổng Dễ + Trung bình + Khó phải bằng số câu của từng chủ đề.")
        seen_topics.add(category)
        topic_rules.append({
            "category": category,
            "count": total,
            "knowledge": {"theory": theory, "practice": practice} if has_knowledge else None,
            "difficulty": {"easy": easy, "medium": medium, "hard": hard} if has_difficulty else None,
        })

    def topic_cell_plan(candidates, rule):
        cells = {
            (knowledge, difficulty): [item for item in candidates if normalize_knowledge(item.get("knowledge_type")) == knowledge and normalize_difficulty(item.get("difficulty")) == difficulty]
            for knowledge in ("theory", "practice")
            for difficulty in ("easy", "medium", "hard")
        }
        plan = {}
        difficulties = ("easy", "medium", "hard")

        def assign(index, theory_left):
            if index == len(difficulties):
                return theory_left == 0
            difficulty = difficulties[index]
            requested = rule["difficulty"][difficulty]
            minimum = max(0, requested - len(cells[("practice", difficulty)]))
            maximum = min(requested, len(cells[("theory", difficulty)]), theory_left)
            for theory_count in range(maximum, minimum - 1, -1):
                practice_count = requested - theory_count
                plan[("theory", difficulty)] = theory_count
                plan[("practice", difficulty)] = practice_count
                if assign(index + 1, theory_left - theory_count):
                    return True
            return False

        if not assign(0, rule["knowledge"]["theory"]):
            raise ValueError("Ngân hàng không đủ câu theo tổ hợp Lý thuyết/Thực hành và độ khó đã chọn cho một chủ đề.")
        return plan
    structured_rules = []
    for rule in structure or []:
        try:
            count = int(rule.get("count") or 0)
        except (TypeError, ValueError):
            raise ValueError("Số câu trong cơ cấu đề phải là số nguyên.")
        if count <= 0:
            continue
        rule_type = str(rule.get("type") or "").strip()
        structured_rules.append({
            "category": str(rule.get("category") or "").strip().casefold(),
            "knowledge_type": str(rule.get("knowledge_type") or "").strip().casefold(),
            "type": _question_type(rule_type) if rule_type else "",
            "difficulty": str(rule.get("difficulty") or "").strip().casefold(),
            "count": count,
        })

    def matches_structure_rule(question, rule):
        return (
            (not rule["category"] or str(question.get("category") or "").strip().casefold() == rule["category"])
            and (not rule["knowledge_type"] or str(question.get("knowledge_type") or "").strip().casefold() == rule["knowledge_type"])
            and (not rule["type"] or _question_type(question.get("type") or "") == rule["type"])
            and (not rule["difficulty"] or str(question.get("difficulty") or "").strip().casefold() == rule["difficulty"])
        )

    if topic_rules and sum(rule["count"] for rule in topic_rules) != questions_per_variant:
        raise ValueError("Tổng số câu của các chủ đề phải bằng số câu mỗi mã đề.")
    for rule in topic_rules:
        available = [question for question in unique_questions if str(question.get("category") or "").strip().casefold() == rule["category"]]
        if len(available) < rule["count"]:
            raise ValueError("Ngân hàng không đủ số câu cho chủ đề đã chọn.")
        if rule["knowledge"] and rule["difficulty"]:
            topic_cell_plan(available, rule)
        elif rule["knowledge"]:
            for knowledge, count in rule["knowledge"].items():
                if len([item for item in available if normalize_knowledge(item.get("knowledge_type")) == knowledge]) < count:
                    raise ValueError("Ngân hàng không đủ câu Lý thuyết/Thực hành cho chủ đề đã chọn.")
    if structured_rules and sum(rule["count"] for rule in structured_rules) != questions_per_variant:
        raise ValueError("Tổng số câu trong cơ cấu loại/kiểu câu hỏi/độ khó phải bằng số câu mỗi mã đề.")
    for rule in structured_rules:
        available = [question for question in unique_questions if matches_structure_rule(question, rule)]
        if len(available) < rule["count"]:
            raise ValueError("Ngân hàng không đủ câu cho một dòng cơ cấu loại/kiểu câu hỏi/độ khó.")

    if difficulty_rule:
        for difficulty, count in difficulty_rule.items():
            available = len([item for item in unique_questions if normalize_difficulty(item.get("difficulty")) == difficulty])
            if available < count:
                raise ValueError(f"Ngan hang khong du {difficulty} cho co cau da chon.")

    seed = int(seed) if seed not in (None, "") else secrets.randbits(63)
    rng = random.Random(seed)
    strata = {}
    for question in unique_questions:
        key = (
            str(question.get("category") or "").strip().casefold(),
            str(question.get("difficulty") or "").strip().casefold(),
            str(question.get("type") or "").strip().casefold(),
            str(question.get("points") or 0),
        )
        strata.setdefault(key, []).append(question)

    quotas = {}
    if not structured_rules:
        total = len(unique_questions)
        raw_targets = {
            key: len(group) * questions_per_variant / total
            for key, group in strata.items()
        }
        quotas = {
            key: min(len(strata[key]), int(raw_targets[key]))
            for key in strata
        }
        while sum(quotas.values()) < questions_per_variant:
            candidates = [key for key in strata if quotas[key] < len(strata[key])]
            if not candidates:
                break
            rng.shuffle(candidates)
            selected_key = max(candidates, key=lambda key: raw_targets[key] - quotas[key])
            quotas[selected_key] += 1

    usage = {str(question["id"]): 0 for question in unique_questions}
    generated = []
    for variant_index in range(1, variant_count + 1):
        selected = []
        selected_ids = set()
        if topic_rules:
            topic_knowledge_plan = {}
            if knowledge_rule and not any(rule["knowledge"] for rule in topic_rules):
                capacities = []
                for rule in topic_rules:
                    category_questions = [question for question in unique_questions if str(question.get("category") or "").strip().casefold() == rule["category"]]
                    capacities.append((
                        len([item for item in category_questions if normalize_knowledge(item.get("knowledge_type")) == "theory"]),
                        len([item for item in category_questions if normalize_knowledge(item.get("knowledge_type")) == "practice"]),
                    ))
                def assign_topic_knowledge(index, theory_left):
                    if index == len(topic_rules):
                        return theory_left == 0
                    total = topic_rules[index]["count"]
                    theory_available, practice_available = capacities[index]
                    minimum = max(0, total - practice_available)
                    maximum = min(total, theory_available, theory_left)
                    for theory_count in range(maximum, minimum - 1, -1):
                        topic_knowledge_plan[index] = theory_count
                        if assign_topic_knowledge(index + 1, theory_left - theory_count):
                            return True
                    return False
                if not assign_topic_knowledge(0, knowledge_rule["theory"]):
                    raise ValueError("Ngân hàng không đủ câu theo cơ cấu Lý thuyết/Thực hành và các chủ đề đã chọn.")
            for rule_index, rule in enumerate(topic_rules):
                candidates = [
                    question for question in unique_questions
                    if str(question["id"]) not in selected_ids
                    and str(question.get("category") or "").strip().casefold() == rule["category"]
                ]
                if rule["knowledge"] and rule["difficulty"]:
                    plan = topic_cell_plan(candidates, rule)
                    for difficulty in ("easy", "medium", "hard"):
                        for knowledge in ("theory", "practice"):
                            count = plan[(knowledge, difficulty)]
                            ranked = [question for question in candidates if str(question["id"]) not in selected_ids and normalize_knowledge(question.get("knowledge_type")) == knowledge and normalize_difficulty(question.get("difficulty")) == difficulty]
                            rng.shuffle(ranked)
                            ranked.sort(key=lambda question: usage[str(question["id"])])
                            if len(ranked) < count:
                                raise ValueError("Ngân hàng không đủ câu không trùng cho cơ cấu chủ đề trong một mã đề.")
                            for question in ranked[:count]:
                                selected.append(question)
                                selected_ids.add(str(question["id"]))
                else:
                    requested = rule["knowledge"]
                    if requested is None and topic_knowledge_plan:
                        theory_count = topic_knowledge_plan[rule_index]
                        requested = {"theory": theory_count, "practice": rule["count"] - theory_count}
                    if requested:
                        for knowledge in ("theory", "practice"):
                            ranked = [question for question in candidates if str(question["id"]) not in selected_ids and normalize_knowledge(question.get("knowledge_type")) == knowledge]
                            rng.shuffle(ranked)
                            ranked.sort(key=lambda question: usage[str(question["id"])])
                            if len(ranked) < requested[knowledge]:
                                raise ValueError("Ngân hàng không đủ câu Lý thuyết/Thực hành không trùng cho chủ đề đã chọn.")
                            for question in ranked[:requested[knowledge]]:
                                selected.append(question)
                                selected_ids.add(str(question["id"]))
                    else:
                        ranked = candidates[:]
                        rng.shuffle(ranked)
                        ranked.sort(key=lambda question: usage[str(question["id"])])
                        for question in ranked[:rule["count"]]:
                            selected.append(question)
                            selected_ids.add(str(question["id"]))
        elif knowledge_rule:
            for knowledge in ("theory", "practice"):
                ranked = [question for question in unique_questions if normalize_knowledge(question.get("knowledge_type")) == knowledge]
                rng.shuffle(ranked)
                ranked.sort(key=lambda question: usage[str(question["id"])])
                if len(ranked) < knowledge_rule[knowledge]:
                    raise ValueError("Ngân hàng không đủ câu theo cơ cấu Lý thuyết/Thực hành đã chọn.")
                for question in ranked[:knowledge_rule[knowledge]]:
                    selected.append(question)
                    selected_ids.add(str(question["id"]))
        elif structured_rules:
            for rule in structured_rules:
                ranked = [
                    question for question in unique_questions
                    if str(question["id"]) not in selected_ids
                    and matches_structure_rule(question, rule)
                ]
                rng.shuffle(ranked)
                ranked.sort(key=lambda question: usage[str(question["id"])])
                if len(ranked) < rule["count"]:
                    raise ValueError("Các dòng cơ cấu đang chồng lấn và không đủ câu hỏi không trùng trong một mã đề.")
                for question in ranked[:rule["count"]]:
                    selected.append(question)
                    selected_ids.add(str(question["id"]))
        else:
            for key, group in strata.items():
                ranked = group[:]
                rng.shuffle(ranked)
                ranked.sort(key=lambda question: usage[str(question["id"])])
                for question in ranked[:quotas[key]]:
                    selected.append(question)
                    selected_ids.add(str(question["id"]))
        if len(selected) < questions_per_variant:
            remaining = [
                question for question in unique_questions
                if str(question["id"]) not in selected_ids
            ]
            rng.shuffle(remaining)
            remaining.sort(key=lambda question: usage[str(question["id"])])
            selected.extend(remaining[:questions_per_variant - len(selected)])
        if difficulty_rule:
            selected_counts = Counter(normalize_difficulty(question.get("difficulty")) for question in selected)
            for target_difficulty in ("easy", "medium", "hard"):
                while selected_counts[target_difficulty] < difficulty_rule[target_difficulty]:
                    replacement = None
                    for source_difficulty in ("easy", "medium", "hard"):
                        if selected_counts[source_difficulty] <= difficulty_rule[source_difficulty]:
                            continue
                        for selected_index, selected_question in enumerate(selected):
                            if normalize_difficulty(selected_question.get("difficulty")) != source_difficulty:
                                continue
                            candidates = [
                                question for question in unique_questions
                                if str(question["id"]) not in selected_ids
                                and normalize_difficulty(question.get("difficulty")) == target_difficulty
                                and str(question.get("category") or "").strip().casefold() == str(selected_question.get("category") or "").strip().casefold()
                                and normalize_knowledge(question.get("knowledge_type")) == normalize_knowledge(selected_question.get("knowledge_type"))
                            ]
                            if candidates:
                                rng.shuffle(candidates)
                                candidates.sort(key=lambda question: usage[str(question["id"])])
                                replacement = (selected_index, candidates[0])
                                break
                        if replacement:
                            break
                    if not replacement:
                        raise ValueError("Ngan hang khong du cau theo dong thoi chu de, Ly thuyet/Thuc hanh va do kho da chon.")
                    selected_index, next_question = replacement
                    previous_question = selected[selected_index]
                    selected_ids.remove(str(previous_question["id"]))
                    selected_ids.add(str(next_question["id"]))
                    selected[selected_index] = next_question
                    selected_counts[normalize_difficulty(previous_question.get("difficulty"))] -= 1
                    selected_counts[target_difficulty] += 1
        # Keep the knowledge check first, then put hands-on work together at the
        # end of every generated version. Both sections stay randomized.
        theory_questions = [question for question in selected if not _is_practical_question(question)]
        practice_questions = [question for question in selected if _is_practical_question(question)]
        rng.shuffle(theory_questions)
        rng.shuffle(practice_questions)
        selected = theory_questions + practice_questions

        for order, source in enumerate(selected, start=1):
            source_id = str(source["id"])
            usage[source_id] += 1
            item = _shuffle_options(source, rng, order + variant_index - 2)
            knowledge_type = normalize_knowledge(item.get("knowledge_type"))
            if knowledge_type in score_rule:
                item["points"] = score_rule[knowledge_type]
            item["source_question_id"] = source_id
            item["variant"] = f"Đề {variant_index}"
            item["order"] = order
            short_hash = hashlib.md5(source_id.encode()).hexdigest()[:10]
            item["id"] = f"de{variant_index}-{order}-{short_hash}"
            generated.append(item)

    warnings = []
    if duplicate_count:
        warnings.append(f"Đã bỏ {duplicate_count} câu trùng nội dung trong file nguồn.")
    if len(unique_questions) < variant_count * questions_per_variant:
        warnings.append(
            "File nguồn chưa đủ để các mã đề hoàn toàn khác nhau; "
            "hệ thống đã giảm trùng lặp giữa các mã ở mức thấp nhất."
        )
    return {
        "questions": generated,
        "variants": [
            {"name": f"Đề {index}", "question_count": questions_per_variant}
            for index in range(1, variant_count + 1)
        ],
        "question_count": len(generated),
        "source_question_count": len(unique_questions),
        "warnings": warnings,
        "generation_config": {
            "variant_count": variant_count,
            "questions_per_variant": questions_per_variant,
            "source_question_count": len(unique_questions),
            "seed": seed,
            "structure": structured_rules,
            "topic_config": topic_rules,
            "knowledge_config": knowledge_rule or {},
            "score_config": score_rule,
            "difficulty_config": difficulty_rule or {},
        },
    }


def google_sheet_export_url(url):
    source = str(url or "").strip()
    spreadsheet = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", source)
    if spreadsheet:
        return f"https://docs.google.com/spreadsheets/d/{spreadsheet.group(1)}/export?format=xlsx"
    drive_file = re.search(r"(?:/file/d/|[?&]id=)([a-zA-Z0-9_-]+)", source)
    if drive_file:
        return f"https://drive.google.com/uc?export=download&id={drive_file.group(1)}"
    raise ValueError("Đường dẫn Google Sheet hoặc tệp Google Drive không hợp lệ.")


def fetch_google_sheet(url):
    response = requests.get(google_sheet_export_url(url), timeout=(10, 90), allow_redirects=True)
    if response.status_code != 200:
        raise ValueError("Không thể đọc nguồn câu hỏi. Hãy bật quyền xem qua liên kết.")
    content_type = response.headers.get("Content-Type", "")
    if "html" in content_type.lower() or response.content[:64].lower().find(b"<html") >= 0:
        raise ValueError("Nguồn câu hỏi đang yêu cầu đăng nhập hoặc chưa được chia sẻ.")
    return response.content

def variants_for(assessment):
    return sorted({str(item.get("variant") or "Đề 1") for item in assessment.questions}, key=str.casefold)


def public_questions(assessment, variant):
    result = []
    for item in assessment.questions:
        if str(item.get("variant") or "Đề 1") != variant:
            continue
        question = {
            key: item.get(key)
            for key in (
                "id", "question_code", "order", "type", "knowledge_type", "text",
                "options", "points", "required", "image_url", "media_url",
                "media_file_id", "answer_image_url", "category", "difficulty",
            )
        }
        if question.get("type") == "matching":
            question["options"] = _normalized_matching_options(question.get("options") or [])
        result.append(question)
    # Existing assessments are also presented theory-first. This keeps the new
    # experience consistent without rewriting stored question data.
    return sorted(
        result,
        key=lambda item: (
            _is_practical_question(item),
            item.get("order") or 0,
            item.get("id") or "",
        ),
    )


def _answer_text(value):
    if isinstance(value, list):
        return ";".join(str(item).strip() for item in value if str(item).strip())
    if isinstance(value, dict):
        return ";".join(
            f"{key}-{value[key]}"
            for key in sorted(value)
            if value[key] not in (None, "") and key not in {"link", "upload_id", "upload_file_id", "upload_url"}
        )
    return str(value or "").strip()


def grade_attempt(attempt):
    questions = [
        item for item in attempt.assessment.questions
        if str(item.get("variant") or "Đề 1") == attempt.variant
    ]
    score = Decimal("0")
    maximum = Decimal("0")
    manual = False
    for question in questions:
        points = Decimal(str(question.get("points") or 0))
        maximum += points
        answer = attempt.answers.get(str(question.get("id")), "")
        answer_text = _answer_text(answer)
        correct = question.get("correct_answers") or []
        if question.get("type") == "single_choice":
            if correct and answer_text.upper() == str(correct[0]).strip().upper():
                score += points
        elif question.get("type") == "multiple_choice":
            actual_values = {part.strip().upper() for part in re.split(r"[,;|]", answer_text) if part.strip()}
            correct_values = {str(part).strip().upper() for part in correct if str(part).strip()}
            if correct_values and actual_values == correct_values:
                score += points
        elif question.get("type") == "short_answer":
            if correct:
                normalized = _key(answer_text)
                if normalized and normalized in {_key(value) for value in correct}:
                    score += points
            else:
                manual = True
        elif question.get("type") in {"matching", "ordering"}:
            if correct and _key(answer_text) in {_key(value) for value in correct}:
                score += points
            elif not correct:
                manual = True
        else:
            manual = True
    attempt.auto_graded_points = score
    attempt.score = score
    attempt.max_score = maximum
    attempt.manual_grading_required = manual
    return attempt


def _safe_sheet_title(value):
    title = re.sub(r"[\\/*?:\[\]]", "-", str(value or "").strip())
    return title[:100] or "Sheet"


def _safe_drive_name(value, fallback="Thư mục"):
    name = re.sub(r'[\\/:*?"<>|\x00-\x1f]', "-", str(value or "").strip())
    name = re.sub(r"\s+", " ", name).strip(" .-")
    return name[:180] or fallback


def _drive_child_folder(service, parent_id, name):
    safe_name = _safe_drive_name(name)
    query_name = safe_name.replace(chr(39), chr(92) + chr(39))
    result = service.files().list(
        q=f"'{parent_id}' in parents and name = '{query_name}' and mimeType = 'application/vnd.google-apps.folder' and trashed = false",
        fields="files(id,name)",
        pageSize=1,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()
    existing = result.get("files", [])
    if existing:
        return existing[0]["id"]
    created = service.files().create(
        body={
            "name": safe_name,
            "parents": [parent_id],
            "mimeType": "application/vnd.google-apps.folder",
        },
        fields="id",
        supportsAllDrives=True,
    ).execute()
    return created["id"]


def _participant_folder_name(service, parent_id, attempt):
    """Use a stable, human-readable identity from the first upload."""
    full_name = _safe_drive_name(attempt.respondent_name, "Người làm")
    identifier = str(attempt.email or attempt.participant_code or attempt.access_token).strip()
    return _safe_drive_name(f"{full_name} - {identifier}", full_name)


def upload_assessment_file_to_drive(uploaded, assessment, attempt, question_id=""):
    raw = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if not raw:
        raise ValueError("Chưa cấu hình GOOGLE_SERVICE_ACCOUNT_JSON để tải tệp lên Google Drive.")
    info = json.loads(raw)
    info["private_key"] = str(info.get("private_key") or "").replace("\\n", "\n")
    credentials = service_account.Credentials.from_service_account_info(
        info,
        scopes=["https://www.googleapis.com/auth/drive.file"],
    )
    service = build("drive", "v3", credentials=credentials, cache_discovery=False)
    folder_id = assessment.drive_folder_id
    config = assessment.storage_config or {}
    if config.get("create_customer_folder", True):
        customer_name = config.get("customer_folder_name") or (
            assessment.partner.name if assessment.partner else assessment.title
        )
        folder_id = _drive_child_folder(service, folder_id, customer_name)
    if config.get("create_participant_folder", True):
        if attempt.drive_folder_id:
            folder_id = attempt.drive_folder_id
        else:
            participant_name = _participant_folder_name(service, folder_id, attempt)
            folder_id = _drive_child_folder(service, folder_id, participant_name)
            attempt.drive_folder_id = folder_id
            attempt.save(update_fields=["drive_folder_id", "updated_at"])
    uploaded.seek(0)
    media = MediaIoBaseUpload(
        io.BytesIO(uploaded.read()),
        mimetype=str(uploaded.content_type or "application/octet-stream"),
        resumable=False,
    )
    result = service.files().create(
        body={"name": _safe_drive_name(f"{question_id} - {uploaded.name}", uploaded.name), "parents": [folder_id]},
        media_body=media,
        fields="id,webViewLink",
        supportsAllDrives=True,
    ).execute()
    return {"id": result.get("id", ""), "url": result.get("webViewLink", "")}


def _variant_sheet_title(prefix, variant):
    label = str(variant or "").strip()
    if _key(label).startswith('de'):
        label = label.split(maxsplit=1)[1] if ' ' in label else label
    return _safe_sheet_title(f"{prefix} {label}".strip())


def _assessment_output_layout(assessment):
    overview_title = "T\u1ed4NG QUAN"
    distribution_title = "PH\u00c2N \u0110\u1ec0"
    delete_log_title = "NH\u1eacT K\u00dd X\u00d3A"
    variants = variants_for(assessment)
    return {
        "overview": overview_title,
        "distribution": distribution_title,
        "delete_log": delete_log_title,
        "question_sheets": {variant: _variant_sheet_title("\u0110\u1ec0", variant) for variant in variants},
        "answer_sheets": {variant: _variant_sheet_title("B\u00c0I L\u00c0M", variant) for variant in variants},
    }


def _sheet_question_header(question, index):
    number = question.get("order") or index
    code = str(question.get("question_code") or "").strip()
    question_text = re.sub(r"\s+", " ", str(question.get("text") or "").strip())
    if len(question_text) > 140:
        question_text = f"{question_text[:137].rstrip()}..."
    kind = "Thực hành · link/tệp riêng" if _is_practical_question(question) else "Câu trả lời"
    identifier = f"{code} · " if code else ""
    return f"Câu {number} · {kind} · {identifier}{question_text or 'Không có nội dung'}"


def _sheet_answer_value(question, value):
    if _is_practical_question(question) and isinstance(value, dict):
        product_link = str(value.get("link") or "").strip()
        evidence_link = str(value.get("upload_url") or "").strip()
        values = []
        if product_link:
            values.append(f"Link sản phẩm: {product_link}")
        if evidence_link:
            values.append(f"Tệp minh chứng: {evidence_link}")
        return "\n".join(values)
    return _answer_text(value)


def _sheet_attempt_row(attempt, questions):
    status_names = {"in_progress": "Đang làm", "submitted": "Đã nộp", "timed_out": "Hết giờ"}
    answer_values = [
        _sheet_answer_value(question, attempt.answers.get(str(question.get("id")), ""))
        for question in questions
    ]
    return [
        attempt.respondent_name, attempt.email, attempt.phone, attempt.organization, attempt.position,
        attempt.participant_code, str(attempt.access_token), attempt.variant,
        status_names.get(attempt.status, attempt.status),
        attempt.started_at.isoformat() if attempt.started_at else "",
        attempt.submitted_at.isoformat() if attempt.submitted_at else "",
        *answer_values,
        float(attempt.auto_graded_points or 0),
        float(attempt.practical_score or 0) if attempt.practical_score is not None else "",
        float(attempt.score or 0),
        "Cần chấm" if attempt.manual_grading_required else "Đã chấm",
        "Đã ghi",
        attempt.synced_at.isoformat() if attempt.synced_at else "",
        attempt.purge_after.isoformat() if attempt.purge_after else "",
    ]


def prepare_assessment_google_sheet(assessment):
    spreadsheet_id = extract_spreadsheet_id(assessment.output_sheet_url)
    if not spreadsheet_id:
        raise ValueError("Chưa cấu hình Google Sheet đầu ra hợp lệ cho đợt kiểm tra.")
    service = build_sheets_service(None, {})
    layout = _assessment_output_layout(assessment)
    metadata = service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets(properties(sheetId,title))",
    ).execute()
    existing = {item.get("properties", {}).get("title") for item in metadata.get("sheets", [])}
    required = [layout["overview"], layout["distribution"], *layout["question_sheets"].values(), *layout["answer_sheets"].values(), layout["delete_log"]]
    requests_body = [
        {"addSheet": {"properties": {"title": title, "gridProperties": {"rowCount": 1000, "columnCount": 250}}}}
        for title in required if title not in existing
    ]
    if requests_body:
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": requests_body},
        ).execute()

    overview_values = [
        ["KI\u1ec2M TRA CU\u1ed0I KH\u00d3A T\u1eacP HU\u1ea4N", assessment.title],
        ["Kh\u00e1ch h\u00e0ng", assessment.partner.name if assessment.partner else ""],
        ["Nh\u00f3m \u0111\u1ed1i t\u01b0\u1ee3ng", assessment.audience_group],
        ["S\u1ed1 ng\u01b0\u1eddi", len(assessment.participants or [])],
        ["S\u1ed1 m\u00e3 \u0111\u1ec1", len(variants_for(assessment))],
        ["Tr\u1ea1ng th\u00e1i", assessment.status],
    ]
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"'{layout['overview']}'!A1",
        valueInputOption="RAW",
        body={"values": overview_values},
    ).execute()

    attempts = list(assessment.attempts.order_by("started_at"))
    attempt_statuses = {}
    participant_identities = set()
    status_names = {"in_progress": "Đang làm", "submitted": "Đã nộp", "timed_out": "Hết giờ"}
    participant_values = [["Họ tên", "Email", "Số điện thoại", "Tổ chuyên môn/Phòng ban", "Chức vụ", "Mã người làm", "Mã lượt làm", "Mã đề", "Trạng thái", "Bắt đầu lúc", "Nộp lúc"]]
    for item in attempts:
        code = str(item.participant_code or "").strip().casefold()
        email = str(item.email or "").strip().casefold()
        phone = str(item.phone or "").strip()
        participant_identities.update(identity for identity in (f"code:{code}" if code else "", f"email:{email}" if email else "", f"phone:{phone}" if phone else "") if identity)
        if code:
            attempt_statuses[code] = item.status or ""
        participant_values.append([
            item.respondent_name, item.email, item.phone, item.organization, item.position,
            item.participant_code, str(item.access_token), item.variant,
            status_names.get(item.status, item.status),
            item.started_at.isoformat() if item.started_at else "",
            item.submitted_at.isoformat() if item.submitted_at else "",
        ])
    participant_values.extend([
        [
            item.get("name", ""), item.get("email", ""), item.get("phone", ""),
            item.get("organization", item.get("group", "")), item.get("position", ""),
            item.get("code", ""), "", item.get("variant", ""),
            status_names.get(attempt_statuses.get(str(item.get("code") or "").strip().casefold(), ""), "Chưa làm"), "", "",
        ]
        for item in assessment.participants or []
        if not any(identity and identity in participant_identities for identity in (
            f"code:{str(item.get('code') or '').strip().casefold()}" if item.get("code") else "",
            f"email:{str(item.get('email') or '').strip().casefold()}" if item.get("email") else "",
            f"phone:{str(item.get('phone') or '').strip()}" if item.get("phone") else "",
        ))
    ])
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"'{layout['distribution']}'!A1",
        valueInputOption="RAW",
        body={"values": participant_values},
    ).execute()

    question_headers = ["STT", "M\u00e3 c\u00e2u h\u1ecfi", "Ch\u1ee7 \u0111\u1ec1", "Lo\u1ea1i c\u00e2u h\u1ecfi", "Ki\u1ec3u c\u00e2u h\u1ecfi", "\u0110\u1ed9 kh\u00f3", "C\u00e2u h\u1ecfi", "\u1ea2nh/Video minh h\u1ecda", "Ph\u01b0\u01a1ng \u00e1n 1", "Ph\u01b0\u01a1ng \u00e1n 2", "Ph\u01b0\u01a1ng \u00e1n 3", "Ph\u01b0\u01a1ng \u00e1n 4", "Ph\u01b0\u01a1ng \u00e1n 5", "\u0110\u00e1p \u00e1n", "\u1ea2nh \u0111\u00e1p \u00e1n", "\u0110i\u1ec3m"]
    for variant, sheet_title in layout["question_sheets"].items():
        rows = [question_headers]
        for question in sorted([item for item in assessment.questions if str(item.get("variant") or "") == variant], key=lambda item: item.get("order") or 0):
            option_map = {str(item.get("key")): item.get("text", "") for item in question.get("options") or []}
            rows.append([
                question.get("order", ""), question.get("question_code", ""), question.get("category", ""), question.get("knowledge_type", ""), question.get("type", ""), question.get("difficulty", ""), question.get("text", ""), question.get("media_url", ""),
                *[option_map.get(str(index), option_map.get(chr(64 + index), "")) for index in range(1, 6)],
                ";".join(str(item) for item in question.get("correct_answers") or []), question.get("answer_image_url", ""), question.get("points", 0),
            ])
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_title}'!A1",
            valueInputOption="RAW",
            body={"values": rows},
        ).execute()

    for variant, sheet_title in layout["answer_sheets"].items():
        questions = public_questions(assessment, variant)
        headers = ["Họ tên", "Email", "Số điện thoại", "Tổ chuyên môn/Phòng ban", "Chức vụ", "Mã người làm", "Mã lượt làm", "Mã đề", "Trạng thái", "Bắt đầu lúc", "Nộp lúc", *[_sheet_question_header(question, index) for index, question in enumerate(questions, start=1)], "Điểm tự động", "Điểm thực hành", "Tổng điểm", "Trạng thái chấm", "Trạng thái đồng bộ", "Thời điểm ghi Sheet", "Hạn xóa dữ liệu tạm"]
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_title}'!A1",
            valueInputOption="RAW",
            body={"values": [headers]},
        ).execute()
    deletion_headers = [["M\u00e3 l\u01b0\u1ee3t l\u00e0m", "M\u00e3 \u0111\u1ec1", "H\u00ecnh th\u1ee9c", "Ng\u01b0\u1eddi y\u00eau c\u1ea7u", "Vai tr\u00f2", "K\u1ebft qu\u1ea3 quy\u1ec1n", "Tr\u1ea1ng th\u00e1i", "Ghi ch\u00fa"]]
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"'{layout['delete_log']}'!A1",
        valueInputOption="RAW",
        body={"values": deletion_headers},
    ).execute()
    return service, spreadsheet_id, layout


def rebuild_assessment_google_sheet_rows(assessment, resources=None):
    service, spreadsheet_id, layout = resources or prepare_assessment_google_sheet(assessment)
    for variant, sheet_title in layout["answer_sheets"].items():
        questions = public_questions(assessment, variant)
        attempts = assessment.attempts.filter(variant=variant).order_by("started_at")
        rows = [_sheet_attempt_row(attempt, questions) for attempt in attempts]
        if rows:
            service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range=f"'{sheet_title}'!A2",
                valueInputOption="RAW",
                body={"values": rows},
            ).execute()
    return service, spreadsheet_id, layout


def append_assessment_deletion_log(attempt, actor, mode, note=""):
    service, spreadsheet_id, layout = prepare_assessment_google_sheet(attempt.assessment)
    service.spreadsheets().values().append(
        spreadsheetId=spreadsheet_id,
        range=f"'{layout['delete_log']}'!A:A",
        valueInputOption="RAW",
        insertDataOption="INSERT_ROWS",
        body={"values": [[
            str(attempt.access_token), attempt.variant, mode, actor, "Admin/System",
            "Hợp lệ", "Đã xóa", note,
        ]]},
    ).execute()


def sync_attempt_to_google_sheet(attempt, resources=None):
    service, spreadsheet_id, layout = resources or prepare_assessment_google_sheet(attempt.assessment)
    questions = public_questions(attempt.assessment, attempt.variant)
    row = _sheet_attempt_row(attempt, questions)
    sheet_title = layout["answer_sheets"][attempt.variant]
    existing = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=f"'{sheet_title}'!G:G",
    ).execute().get("values", [])
    row_number = next(
        (index for index, values in enumerate(existing, start=1) if values and str(values[0]) == str(attempt.access_token)),
        None,
    )
    if row_number:
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_title}'!A{row_number}",
            valueInputOption="RAW",
            body={"values": [row]},
        ).execute()
    else:
        service.spreadsheets().values().append(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_title}'!A:A",
            valueInputOption="RAW",
            insertDataOption="INSERT_ROWS",
            body={"values": [row]},
        ).execute()
    return True
